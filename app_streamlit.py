import streamlit as st
import os
import tempfile
import pandas as pd
import json
import re
import unicodedata
import io
import requests
from openai import OpenAI
from contextlib import contextmanager
import warnings
import logging
import sys

# Configurações da página
st.set_page_config(
    page_title="PDFator - Extrator de Faturas de Energia",
    page_icon="⚡",
    layout="wide"
)

# Ocultar avisos e mensagens de debug
warnings.filterwarnings("ignore")
logging.getLogger().setLevel(logging.ERROR)

# Função para ocultar saída
@contextmanager
def ocultar_output():
    old_stdout = sys.stdout
    old_stderr = sys.stderr
    sys.stdout = io.StringIO()
    sys.stderr = io.StringIO()
    try:
        yield
    finally:
        sys.stdout = old_stdout
        sys.stderr = old_stderr

# Estilo personalizado para uma interface limpa
st.markdown("""
<style>
    .main {
        background-color: #f8f9fa;
    }
    .stProgress > div > div {
        background-color: #4CAF50;
    }
    .stExpander {
        border: 1px solid #e6e6e6;
        border-radius: 5px;
    }
</style>
""", unsafe_allow_html=True)

# Título e cabeçalho
st.title("⚡ PDFator - Extrator de Faturas de Energia")
st.write("Extraia e categorize informações de suas faturas de energia em poucos cliques.")

# Inicializar variáveis de sessão
if "processed_files" not in st.session_state:
    st.session_state.processed_files = []
if "current_file_index" not in st.session_state:
    st.session_state.current_file_index = 0
if "total_cost" not in st.session_state:
    st.session_state.total_cost = 0.0
if "api_key" not in st.session_state:
    st.session_state.api_key = ""
if "processing_complete" not in st.session_state:
    st.session_state.processing_complete = False

# Mapeamento de tipos de cobrança
mapeamento_tipo_cobranca = {
    # Encargo
    "Energia elet. adquirida 3": "Encargo",
    "energia elet adquirida 3": "Encargo",
    "energiaelet.adquirida3": "Encargo",
    "energia elet. adquirida3": "Encargo",
    "energia elet. adquirida 3* c/imposto": "Encargo",
    "energia elet. adquirida 3* c/lmposto": "Encargo",
    "energia elet. adquirida 3* c imp": "Encargo",
    "energia elet adquirida 3 c imposto": "Encargo",
    "Liminar Red Icms Fatura": "Encargo",
    "liminar red icms fatura": "Encargo",
    "liminarredicmsfatura": "Encargo",
    "liminar red. icms fatura": "Encargo",
    "Subtotal Faturamento": "Encargo",
    "Componente Encargo kWh HFP": "Encargo",
    "componente encargo kwh hfp": "Encargo",
    "Componente Encargo kWh HP": "Encargo",
    "componente encargo kwh hp": "Encargo",
    "Energia Terc Comercializad HFP": "Encargo",
    "energia terc comercializad hfp": "Encargo",
    "Energia Terc Comercializad HP": "Encargo",
    "energia terc comercializad hp": "Encargo",
    "Energia Isenta Comercial. HFP": "Encargo",
    "energia isenta comercial. hfp": "Encargo",
    "PIS/COFINS da subvenção/descon": "Encargo",
    "pis/cofins da subvenção/descon": "Encargo",
    "subtotal faturamento": "Encargo",
    "sub total faturamento": "Encargo",
    "subtot. faturamento": "Encargo",
    "Componente Encargo kWh HFP": "Encargo",
    "componente encargo kwh hfp": "Encargo",
    "componente encargokwh hfp": "Encargo",
    "componente encargo kw hfp": "Encargo",
    "Componente Encargo kWh HP": "Encargo",
    "componente encargo kwh hp": "Encargo",
    "componente encargokwh hp": "Encargo",
    "componente encargo kw hp": "Encargo",
    "Componente Encargo HFP": "Encargo",
    "componente encargo hfp": "Encargo",
    "Componente Encargo HP": "Encargo",
    "componente encargo hp": "Encargo",
    "ENERGIA ACL": "Encargo",
    "energia acl": "Encargo",
    "energiaacl": "Encargo",
    "energia acl ": "Encargo",
    "CONSUMOATIVO PONTA TUSD": "Encargo",
    "consumoativo ponta tusd": "Encargo",
    "consumo ativo ponta tusd": "Encargo",
    "consumo ativoponta tusd": "Encargo",
    "CONSUMO ATIVO F.PONTA TUSD": "Encargo",
    "consumo ativo f.ponta tusd": "Encargo",
    "consumo ativo f. ponta tusd": "Encargo",
    "consumo ativo f ponta tusd": "Encargo",
    "BENEFICIO TARIFARIO BRUTO": "Encargo",
    "beneficio tarifario bruto": "Encargo",
    "benefício tarifario bruto": "Encargo",
    "beneficio tar. bruto": "Encargo",
    "benefício tar. bruto": "Encargo",
    "PIS/COFINS da subvencao/descon": "Encargo",
    "PIS/COFINS da subvencäo/descon": "Encargo",
    "pis cofins da subvençao/descon": "Encargo",
    "piscofins da subvencao descon": "Encargo",
    "Correcao IPCA/IGPM s/ conta 01/25 pg 19/02/25": "Encargo",
    "correção ipca igpm s/ conta": "Encargo",
    "correcao ipca igpm s/conta": "Encargo",
    "correcao ipca/igpm s/conta": "Encargo",
    "correcao ipca igpm s conta": "Encargo",
    "Componente Encargo kWh HFP": "Encargo",
    "componente encargo kwh hfp": "Encargo",
    "componente encargokwh hfp": "Encargo",
    "Componente Encargo kWh HP": "Encargo",
    "componente encargo kwh hp": "Encargo",
    "componente encargokwh hp": "Encargo",
    "PIS/COFINS da subvencao/descon": "Encargo",
    "PIS/COFINS da subvencäo/descon": "Encargo",
    "PIS/COFINS da subvenção/descon": "Encargo",
    "pis cofins da subvençao descon": "Encargo",
    "Encargo Cta Covid REN 885/2020": "Encargo",
    "encargo cta covid ren885/2020": "Encargo",
    "encargo covid ren 885/2020": "Encargo",
    "encargo covid ren885/2020": "Encargo",
    "Energia Ativa kWh HFP/nico": "Encargo",
    "energia ativa kwh hfp/unico": "Encargo",
    "energia ativa kwh hfp unico": "Encargo",
    "energia ativa kwh hfp": "Encargo",
    "Energia Ativa kWh HP": "Encargo",
    "energia ativa kwh hp": "Encargo",
    "DEBITO VAR IPCA": "Encargo",
    "débito var ipca": "Encargo",
    "debito var ipca": "Encargo",
    "deb1to var ipca": "Encargo",
    "ENERGIA ACL": "Encargo",
    "energia acl": "Encargo",
    "energiaacl": "Encargo",
    "TUSD ponta": "Encargo",
    "TUSD fora ponta": "Encargo",
    "tusd ponta": "Encargo",
    "tusd fora ponta": "Encargo",
    "Consumo Energia Elétrica HP": "Encargo",
    "consumo energia eletrica hp": "Encargo",
    "consumo energia elétrica hp": "Encargo",
    "DEMANDALEIESTADUAL16.886/18": "Encargo",

    # Desc. Encargo
    "DEDUCAO ENERGIA ACL": "Desc. Encargo",
    "DEDUCÃO ENERGIA ACL": "Desc. Encargo",
    "DED. ENERGIA ACL": "Desc. Encargo",
    "D. ENERGIA ACL": "Desc. Encargo",
    "Valor de Autoprodução HFP": "Desc. Encargo",
    "BENEFICIO TARIFARIO LIQUIDO": "Desc. Encargo",
    "BENEFÍCIO TARIFÁRIO LIQUIDO": "Desc. Encargo",
    "BENEFiCIO TARIFARIO LIQUIDO": "Desc. Encargo",
    "BENEFÍCIO TARIFÁRIO LÍQUIDO": "Desc. Encargo",
    "BENEFICIO TARIFÁRIO LÍQUIDO": "Desc. Encargo",
    # ... Resto do mapeamento para Desc. Encargo

    # Demanda
    "demanda ponta c/desconto": "Demanda",
    "demanda ponta cidesconto": "Demanda",
    "demanda ponta cidescon": "Demanda",
    "demanda ponta c/desc0nt0": "Demanda",
    "DEMANDA FORA PONTA C/DESC": "Demanda",
    "DEMANDA FORA PONTA C/DESC0": "Demanda",
    "DEMANDA FORA PONTA C/DESCONTO": "Demanda",
    "DEMANDA FORA PONTA C/DESC0NT0": "Demanda",
    "DEMANDA FORA PONTA CIDESCONTO": "Demanda",
    "DEMANDA FORA PONTA CIDESC0NTO": "Demanda",
    # ... Resto do mapeamento para Demanda

    # Desc. Demanda
    "DIF.FATUR.TUSD-HOMOLOG.CCEE": "Desc. Demanda",
    "DIF.FATUR.TUSD-HOMOLOG.C0EE": "Desc. Demanda",
    "DIF.FATUR.TUSD-HOMOLOG.0CEE": "Desc. Demanda",
    "DIF.FATUR.TUSD HOMOLOG CCEE": "Desc. Demanda",
    "DIF.FATUR TUSD-HOMOLOG CCEE": "Desc. Demanda",
    # ... Resto do mapeamento para Desc. Demanda

    # Reativa
    "UFER PONTA TE": "Reativa",
    "UFERFORA PONTA TE": "Reativa",
    "UFER PONTA  TE": "Reativa",
    "UFER PONTA TE ": "Reativa",
    "UFERF0RA PONTA TE": "Reativa",
    "UFER P0NTA TE": "Reativa",
    # ... Resto do mapeamento para Reativa

    # Contb. Publica
    "CIP-ILUM PUB PREF MUNICIPAL": "Contb. Publica",
    "CIP - ILUM PUB PREF MUNICIPAL": "Contb. Publica",
    "COSIP.SAOPAULO-MUNICIPAL": "Contb. Publica",
    "CIP-BARUERI-MUNICIPAL": "Contb. Publica",
    "CIP-CAJAMAR-MUNICIPAL": "Contb. Publica",
    "CIP-ILUM PUB PREF MUNICIPALCONTRIB PUB": "Contb. Publica",
    # ... Resto do mapeamento para Contb. Publica
}

# Funções de utilidade
def normalizar_texto(texto):
    if not isinstance(texto, str):
        return ""
    texto = texto.strip().lower()
    texto = unicodedata.normalize('NFKD', texto).encode('ASCII', 'ignore').decode('ASCII')
    texto = re.sub(r'[^a-z0-9 ]', '', texto)  # remove qualquer coisa que não for letra, número ou espaço
    texto = " ".join(texto.split())  # remove espaços duplicados
    return texto

def obter_cotacao_dolar():
    try:
        url = "https://economia.awesomeapi.com.br/json/last/USD-BRL"
        resp = requests.get(url, timeout=5)
        if resp.status_code == 200:
            data = resp.json()
            return float(data["USDBRL"]["bid"])
        else:
            return 5.0  # Valor padrão se não conseguir obter
    except Exception as e:
        return 5.0  # Valor padrão em caso de erro

def calcular_custo_api(response):
    PRECO_PROMPT = 0.0005
    PRECO_COMPLETION = 0.0015
    usage = response.usage
    pt = usage.prompt_tokens
    ct = usage.completion_tokens
    custo_usd = (pt / 1000) * PRECO_PROMPT + (ct / 1000) * PRECO_COMPLETION
    return custo_usd

def corrigir_negativo_final(valor):
    if isinstance(valor, str) and valor.strip().endswith('-'):
        valor_corrigido = '-' + valor.strip()[:-1]
        return valor_corrigido
    return valor

def formatar_numero_brasileiro(x):
    try:
        num = float(str(x).replace(",", "."))
        return f"{num:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except:
        return x

def extrair_texto(arquivo_pdf):
    try:
        import fitz  # PyMuPDF
        import io
        from PIL import Image
        
        texto = ""
        
        # Criar arquivo temporário
        with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as temp_file:
            temp_file.write(arquivo_pdf.read())
            temp_path = temp_file.name
            
        # Extrair texto do PDF
        try:
            doc = fitz.open(temp_path)
            for pagina in doc:
                texto += pagina.get_text()
            doc.close()
        except Exception as e:
            st.error(f"⚠️ Erro ao ler o PDF: {str(e)}")
            
        # Limpar arquivo temporário
        os.unlink(temp_path)
        
        return texto
    except Exception as e:
        st.error(f"⚠️ Erro ao processar o PDF: {str(e)}")
        return ""

def extrair_trecho_itens_fatura(texto):
    texto = texto.replace("\\n", "\n")
    inicio = texto.find("Itens da Fatura")
    if inicio == -1:
        inicio = texto.find("Itens")
        if inicio == -1:
            return ""
    trecho = texto[inicio:]
    palavras_fim = ["Resumo", "Histórico de Consumo", "Medições", "DADOS DE MEDIÇÃO"]
    fim = len(trecho)
    for p in palavras_fim:
        idx = trecho.find(p)
        if 0 < idx < fim:
            fim = idx
    return trecho[:fim]

def extrair_info_cabecalho(texto_pdf, client):
    # Limitar a 500 palavras para melhorar a performance
    texto_limitado = " ".join(texto_pdf.split()[:150])

    prompt_extracao_cabecalho = """
    Você é um assistente especializado em extrair informações de documentos.

    Extraia APENAS as seguintes informações da conta de energia elétrica:
    1. Número de Instalação (pode aparecer como "Nº da Instalação", "Instalação", "Nº Instalação", "Unid. Consumidora", "UC", "Código do Cliente", etc.)
    2. Referência (geralmente um mês/ano, como "01/2025", "JAN/2025", "JANEIRO/2025", etc.)
    3. A REFERENCIA SEMPRE É UM VALOR NUMÉRICO!!!

    Retorne apenas um objeto JSON com os seguintes campos:
    {
        "N_Instalacao": "valor extraído",
        "Referencia": "valor extraído"
    }

    Se alguma informação não for encontrada, use string vazia.
    Não inclua explicações, apenas o JSON.
    """

    try:
        response = client.chat.completions.create(
            model="gpt-3.5-turbo-0125",
            messages=[
                {"role": "system", "content": prompt_extracao_cabecalho},
                {"role": "user", "content": texto_limitado}
            ],
            temperature=0
        )

        custo = calcular_custo_api(response)
        resposta = response.choices[0].message.content.strip()

        # Limpar qualquer formatação extra
        if resposta.startswith("```json"):
            resposta = resposta.replace("```json", "").replace("```", "").strip()

        try:
            dados = json.loads(resposta)
            return dados, custo
        except:
            # Se falhar ao fazer parse do JSON, retorna valores vazios
            return {"N_Instalacao": "", "Referencia": ""}, custo

    except Exception as e:
        st.error(f"⚠️ Erro ao extrair informações de cabeçalho: {str(e)}")
        return {"N_Instalacao": "", "Referencia": ""}, 0

def processar_pdf(arquivo_pdf, client, progress_bar, status_text):
    try:
        # Extrair texto do PDF
        texto_completo = extrair_texto(arquivo_pdf)
        
        if not texto_completo:
            st.error(f"⚠️ Não foi possível extrair texto do arquivo {arquivo_pdf.name}")
            return None
        
        # Extrair informações de cabeçalho (N. Instalação e Referência)
        info_cabecalho, custo_cabecalho = extrair_info_cabecalho(texto_completo, client)
        
        # Extrair trecho específico com os itens da fatura
        texto_tabela = extrair_trecho_itens_fatura(texto_completo)
        
        if not texto_tabela.strip():
            st.error(f"⚠️ Seção 'Itens da Fatura' não encontrada no arquivo {arquivo_pdf.name}")
            return None
        
        # Prompt para extração das informações da tabela
        prompt_extracao = """
        Você é um extrator especializado em identificar e estruturar tabelas extraídas de documentos PDF.

        Instruções:
        - Você receberá o conteúdo de um trecho de um PDF.
        - Localize a tabela que se inicia com o título "Itens da Fatura" e processe até o último valor numérico listado.
        - Cada linha da tabela deve ser transformada em um objeto JSON com os seguintes campos obrigatórios:
          ["Item", "Unidade", "Quantidade", "Preço Unitário (R$)", "Valor (R$)", "PIS/COFINS", "Base de Cálculo ICMS", "Alíquota ICMS", "ICMS (R$)", "Tarifa Unitária"].

        Regras de extração:
        - Ignore qualquer linha de uma estrutracao de tabela que comece com "Medidor" e nao deixe isso afetar a extracao para outras linhas.
        - A linha que contém o texto "TOTAL" deve ser separada como um objeto independente.
        - Não mescle o conteúdo da linha "TOTAL" com nenhuma outra linha da tabela.
        - Se uma linha contiver apenas um nome e números, associe o maior número ao campo "Valor (R$)".
        - Não infira (não invente) valores ou informações ausentes.
        - Campos vazios devem ser preenchidos com a string vazia: `""`.
        - Extraia corretamente números negativos, mesmo quando o sinal de negativo estiver **posicionado após o número** (exemplo: 32123,23- deve ser interpretado como -32123,23).
        - É essencial que você extraia TODAS as linhas da tabela.
        Regra especial:
        - Se encontrar a linha "Contrib Ilum Publica Municipal":
            - Preencher apenas o campo "Valor (R$)" com o valor encontrado.
            - Deixar os campos "Unidade", "Quantidade", "Preço Unitário (R$)", "PIS/COFINS", "Base de Cálculo ICMS", "Alíquota ICMS", "ICMS (R$)", "Tarifa Unitária" como string vazia `""`.

        Caso a tabela não seja encontrada:
        - Retorne exatamente: ERRO: tabela nao encontrada.

        Importante:
        - Seja extremamente rigoroso na separação de colunas e na associação dos números aos campos corretos.
        """
        
        # Fazer a requisição para a API da OpenAI
        response = client.chat.completions.create(
            model="gpt-3.5-turbo-0125",
            messages=[
                {"role": "system", "content": prompt_extracao},
                {"role": "user", "content": texto_tabela}
            ],
            temperature=0,
            max_tokens=4096
        )
        
        custo_api = calcular_custo_api(response)
        custo_total = custo_cabecalho + custo_api
        
        resposta = response.choices[0].message.content.strip()
        if resposta.startswith("```json"):
            resposta = resposta.replace("```json", "").replace("```", "").strip()
        
        if resposta.lower().startswith("erro"):
            st.error(f"⚠️ {resposta}")
            return None
        
        # Converter a resposta para um DataFrame
        try:
            tabela = pd.DataFrame(json.loads(resposta))
        except Exception as e:
            st.error(f"⚠️ Erro ao processar tabela: {str(e)}")
            return None
        
        # Calcular resumo por tipo de cobrança
        resumo = {"Encargo": 0, "Desc. Encargo": 0, "Demanda": 0, "Desc. Demanda": 0, "Reativa": 0, "Contb. Publica": 0, "Outros": 0}
        
        for _, row in tabela.iterrows():
            item = str(row["Item"]).strip().lower()
            valor_corrigido = corrigir_negativo_final(str(row["Valor (R$)"]))
            valor = pd.to_numeric(valor_corrigido.replace(".", "").replace(",", "."), errors='coerce')
            
            if pd.isna(valor) or item == "":
                continue
            
            if any(termo in item for termo in ["total", "subtotal faturamento", "subtotal outros"]):
                continue
            
            tipo = "Outros"
            for chave in mapeamento_tipo_cobranca:
                if normalizar_texto(chave) in normalizar_texto(item) or normalizar_texto(item) in normalizar_texto(chave):
                    tipo = mapeamento_tipo_cobranca[chave]
                    break
            
            resumo[tipo] += valor
        
        # Formatar números nas colunas apropriadas
        colunas_numericas = ["Quantidade", "Preço Unitário (R$)", "Valor (R$)", "PIS/COFINS", 
                            "Base de Cálculo ICMS", "Alíquota ICMS", "ICMS (R$)", "Tarifa Unitária"]
        
        for coluna in colunas_numericas:
            if coluna in tabela.columns:
                tabela[coluna] = tabela[coluna].apply(corrigir_negativo_final).apply(formatar_numero_brasileiro)
        
        # Retornar resultados
        return {
            "nome_arquivo": arquivo_pdf.name,
            "tabela": tabela,
            "resumo": resumo,
            "info_cabecalho": info_cabecalho,
            "custo": custo_total
        }
            
    except Exception as e:
        st.error(f"⚠️ Erro ao processar o arquivo {arquivo_pdf.name}: {str(e)}")
        return None

def exportar_excel(resultados):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
        
        # Inicializar novo Excel
        wb = openpyxl.Workbook()
        
        # Criar aba de resumo
        if 'Sheet' in wb.sheetnames:
            ws_resumo = wb['Sheet']
            ws_resumo.title = "Resumo"
        else:
            ws_resumo = wb.create_sheet(title="Resumo")
        
        ws_resumo.sheet_view.showGridLines = False
        
        # Adicionar cabeçalho na aba de resumo
        cabecalhos_resumo = ["Nome Do PDF", "N Instalacao", "Referencia", "Encargo", "Desc. Encargo", 
                            "Reativa", "Demanda", "Desc. Demanda", "Contrib. Publica"]
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        
        for col, header in enumerate(cabecalhos_resumo, start=1):
            cell = ws_resumo.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # Adicionar dados na aba de resumo
        for i, resultado in enumerate(resultados, start=2):
            if not resultado:  # Pular se resultado for None
                continue
                
            ws_resumo.cell(row=i, column=1, value=resultado["nome_arquivo"])
            ws_resumo.cell(row=i, column=2, value=resultado["info_cabecalho"].get("N_Instalacao", ""))
            ws_resumo.cell(row=i, column=3, value=resultado["info_cabecalho"].get("Referencia", ""))
            
            # Adicionar valores do resumo
            ws_resumo.cell(row=i, column=4, value=formatar_numero_brasileiro(resultado["resumo"].get("Encargo", 0)))
            ws_resumo.cell(row=i, column=5, value=formatar_numero_brasileiro(resultado["resumo"].get("Desc. Encargo", 0)))
            ws_resumo.cell(row=i, column=6, value=formatar_numero_brasileiro(resultado["resumo"].get("Reativa", 0)))
            ws_resumo.cell(row=i, column=7, value=formatar_numero_brasileiro(resultado["resumo"].get("Demanda", 0)))
            ws_resumo.cell(row=i, column=8, value=formatar_numero_brasileiro(resultado["resumo"].get("Desc. Demanda", 0)))
            ws_resumo.cell(row=i, column=9, value=formatar_numero_brasileiro(resultado["resumo"].get("Contb. Publica", 0)))
        
        # Criar abas individuais para cada PDF
        for resultado in resultados:
            if not resultado:  # Pular se resultado for None
                continue
                
            nome_aba = resultado["nome_arquivo"][:31]  # Excel limita a 31 caracteres
            ws = wb.create_sheet(title=nome_aba)
            ws.sheet_view.showGridLines = False
            
            tabela = resultado["tabela"]
            
            start_row = 3
            start_col = 2
            
            # Adicionar cabeçalho da tabela
            for j, col_name in enumerate(tabela.columns, start=start_col):
                cell = ws.cell(row=start_row, column=j, value=col_name)
                cell.font = header_font
                cell.fill = header_fill
            
            # Adicionar dados da tabela
            for i, row in enumerate(tabela.values, start=start_row + 1):
                for j, value in enumerate(row, start=start_col):
                    ws.cell(row=i, column=j, value=value)
            
            # Ajustar largura das colunas
            for col in ws.columns:
                max_length = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = max(max_length + 2, 10)
            
            # Adicionar resumo na aba do PDF
            resumo_row_start = start_row + len(tabela) + 5
            
            ws.cell(row=resumo_row_start, column=start_col, value="Tipo de Cobrança").font = header_font
            ws.cell(row=resumo_row_start, column=start_col, value="Tipo de Cobrança").fill = header_fill
            ws.cell(row=resumo_row_start, column=start_col + 1, value="Valor Total (R$)").font = header_font
            ws.cell(row=resumo_row_start, column=start_col + 1, value="Valor Total (R$)").fill = header_fill
            
            tipos = ["Encargo", "Desc. Encargo", "Demanda", "Desc. Demanda", "Reativa", "Contb. Publica", "Outros", "Total"]
            total_resumo = sum(resultado["resumo"].values())
            
            for idx_tipo, tipo in enumerate(tipos):
                valor_tipo = resultado["resumo"].get(tipo, 0) if tipo != "Total" else total_resumo
                ws.cell(row=resumo_row_start + idx_tipo + 1, column=start_col, value=tipo)
                ws.cell(row=resumo_row_start + idx_tipo + 1, column=start_col + 1, value=formatar_numero_brasileiro(round(valor_tipo, 2)))
        
        # Ajustar largura das colunas na aba de resumo
        for col in ws_resumo.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws_resumo.column_dimensions[col_letter].width = max(max_length + 2, 12)
        
        # Remover a aba Sheet padrão se existir e não tiver sido usada
        if 'Sheet' in wb.sheetnames and 'Sheet' != ws_resumo.title:
            wb.remove(wb['Sheet'])
        
        # Salvar temporariamente
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
        
    except Exception as e:
        st.error(f"⚠️ Erro ao exportar para Excel: {str(e)}")
        return None

# Função para navegar entre os arquivos processados
def navegar_anterior():
    if st.session_state.current_file_index > 0:
        st.session_state.current_file_index -= 1

def navegar_proximo():
    if st.session_state.current_file_index < len(st.session_state.processed_files) - 1:
        st.session_state.current_file_index += 1

# Interface principal
col1, col2 = st.columns([2, 1])

with col1:
    # Campo seguro para a chave da API da OpenAI
    api_key = st.text_input("🔑 Insira sua chave da API OpenAI:", 
                           type="password", 
                           help="Sua chave não será armazenada permanentemente")
    
    if api_key:
        st.session_state.api_key = api_key

with col2:
    # Contagem de PDFs e custo estimado
    st.metric("📄 PDFs processados", len(st.session_state.processed_files))
    
    cotacao = obter_cotacao_dolar()
    custo_brl = st.session_state.total_cost * cotacao
    st.metric("💰 Custo estimado", f"${st.session_state.total_cost:.6f} USD ≈ R${custo_brl:.2f}")

# Upload de arquivos
uploaded_files = st.file_uploader("📂 Selecione seus arquivos PDF (máximo 80 arquivos):", 
                                 type=["pdf"], 
                                 accept_multiple_files=True,
                                 help="Upload de até 80 arquivos PDF por vez")

# Limitar a 80 PDFs
if uploaded_files and len(uploaded_files) > 80:
    st.error("⚠️ Limite excedido! Você pode carregar no máximo 80 arquivos PDF por vez.")
    uploaded_files = uploaded_files[:80]

# Botão para processar PDFs
if uploaded_files and st.session_state.api_key and st.button("🔍 Processar PDFs"):
    # Validar se a chave da API é válida
    try:
        client = OpenAI(api_key=st.session_state.api_key)
        
        # Barra de progresso
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        resultados = []
        
        # Processar cada PDF
        for i, uploaded_file in enumerate(uploaded_files):
            status_text.text(f"Processando arquivo {i+1} de {len(uploaded_files)}: {uploaded_file.name}")
            progress_bar.progress((i) / len(uploaded_files))
            
            resultado = processar_pdf(uploaded_file, client, progress_bar, status_text)
            
            if resultado:
                st.session_state.processed_files.append(resultado)
                resultados.append(resultado)
                st.session_state.total_cost += resultado["custo"]
            
            progress_bar.progress((i + 1) / len(uploaded_files))
        
        status_text.text("Gerando arquivo Excel...")
        
        # Exportar para Excel
        excel_file = exportar_excel(resultados)
        
        if excel_file:
            progress_bar.progress(1.0)
            status_text.text("✅ Processamento concluído!")
            st.session_state.processing_complete = True
            
            st.download_button(
                label="📥 Baixar Resultados em Excel",
                data=excel_file,
                file_name="resultados_faturas.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            st.error("⚠️ Não foi possível gerar o arquivo Excel.")
        
    except Exception as e:
        if "Invalid API key" in str(e):
            st.error("⚠️ Chave de API da OpenAI inválida. Por favor, verifique e tente novamente.")
        else:
            st.error(f"⚠️ Erro ao processar os arquivos: {str(e)}")

# Exibir arquivos processados com navegação por setas
if st.session_state.processed_files:
    st.subheader("📑 Arquivos Processados")
    
    total_arquivos = len(st.session_state.processed_files)
    
    # Navegação com setas
    col1, col2, col3 = st.columns([1, 3, 1])
    
    with col1:
        if st.button("⬅️ Anterior", on_click=navegar_anterior, disabled=(st.session_state.current_file_index <= 0)):
            pass
    
    with col2:
        st.write(f"Arquivo {st.session_state.current_file_index + 1} de {total_arquivos}")
    
    with col3:
        if st.button("➡️ Próximo", on_click=navegar_proximo, disabled=(st.session_state.current_file_index >= total_arquivos - 1)):
            pass
    
    # Exibir o PDF atual
    if total_arquivos > 0:
        arquivo = st.session_state.processed_files[st.session_state.current_file_index]
        
        # Mostrar informações de cabeçalho
        st.write(f"**Nome do arquivo:** {arquivo['nome_arquivo']}")
        st.write(f"**Número de Instalação:** {arquivo['info_cabecalho'].get('N_Instalacao', 'N/A')}")
        st.write(f"**Referência:** {arquivo['info_cabecalho'].get('Referencia', 'N/A')}")
        
        # Mostrar resumo dos valores
        st.subheader("Resumo de Valores")
        resumo_df = pd.DataFrame([
            {"Tipo": k, "Valor (R$)": formatar_numero_brasileiro(v)} 
            for k, v in arquivo["resumo"].items()
        ])
        st.dataframe(resumo_df, use_container_width=True)
        
        # Mostrar tabela detalhada
        st.subheader("Detalhes da Fatura")
        st.dataframe(arquivo["tabela"], use_container_width=True)

# Rodapé informativo
st.divider()
st.caption("📝 PDFator - Extrator de Faturas de Energia | 2025")
st.caption("💡 Dica: O custo estimado é baseado no uso da API OpenAI para processamento dos PDFs.")
