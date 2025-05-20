import streamlit as st
import os
import tempfile
import base64
import json
import pandas as pd
import re
import io
import time
from datetime import datetime
import pdfplumber
from PIL import Image, ImageDraw
import PyPDF2
import pikepdf
import requests
from openai import OpenAI
import unicodedata
import warnings
from pdf2image import convert_from_bytes
import numpy as np
import openpyxl
warnings.filterwarnings('ignore')

# Configurações da página
st.set_page_config(
    page_title="Leitor de Faturas",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Estilo CSS personalizado
st.markdown("""
<style>
    .main {
        padding: 1rem;
    }
    .stButton button {
        width: 100%;
        border-radius: 5px;
        height: 3em;
        font-weight: bold;
    }
    .stTextInput>div>div>input {
        border-radius: 5px;
    }
    .stFileUploader>div>button {
        border-radius: 5px;
    }
    .highlight {
        background-color: #f0f2f6;
        padding: 1rem;
        border-radius: 5px;
        margin-bottom: 1rem;
    }
    .cost-display {
        background-color: #e6f3ff;
        padding: 0.5rem;
        border-radius: 5px;
        margin-top: 1rem;
    }
    .error-message {
        color: red;
        font-weight: bold;
    }
    .success-message {
        color: green;
        font-weight: bold;
    }
    .selection-instructions {
        background-color: #f8f9fa;
        padding: 10px;
        border-radius: 5px;
        margin-bottom: 15px;
        border-left: 4px solid #4CAF50;
    }
    .pdf-image {
        border: 1px solid #ddd;
        border-radius: 5px;
        margin-bottom: 10px;
    }
    .selection-area {
        position: absolute;
        border: 2px dashed;
        pointer-events: none;
    }
    .ref-area {
        border-color: blue;
        background-color: rgba(0, 0, 255, 0.1);
    }
    .tab-area {
        border-color: green;
        background-color: rgba(0, 255, 0, 0.1);
    }
    .area-preview {
        margin-top: 10px;
        padding: 10px;
        border-radius: 5px;
        border: 1px solid #ddd;
    }
    .ref-preview {
        background-color: rgba(0, 0, 255, 0.1);
        border-left: 4px solid blue;
    }
    .tab-preview {
        background-color: rgba(0, 255, 0, 0.1);
        border-left: 4px solid green;
    }
    .stSlider > div > div > div {
        background-color: #4CAF50;
    }
    .stSlider > div > div > div[aria-valuemin] {
        background-color: #2196F3;
    }
    .selection-box {
        border: 3px solid;
        position: absolute;
        pointer-events: none;
    }
    .blue-box {
        border-color: blue;
        background-color: rgba(0, 0, 255, 0.2);
    }
    .green-box {
        border-color: green;
        background-color: rgba(0, 255, 0, 0.2);
    }
    .step-indicator {
        background-color: #4CAF50;
        color: white;
        padding: 5px 10px;
        border-radius: 15px;
        margin-right: 10px;
        font-weight: bold;
    }
    .step-text {
        font-weight: bold;
    }
    .step-container {
        display: flex;
        align-items: center;
        margin-bottom: 10px;
    }
    .slider-container {
        background-color: #f8f9fa;
        padding: 15px;
        border-radius: 5px;
        margin-bottom: 15px;
    }
    .slider-label {
        font-weight: bold;
        margin-bottom: 5px;
    }
    .slider-blue {
        color: blue;
    }
    .slider-green {
        color: green;
    }
    .apply-button {
        background-color: #4CAF50;
        color: white;
        padding: 10px 15px;
        border: none;
        border-radius: 5px;
        cursor: pointer;
        font-weight: bold;
        margin-top: 10px;
        width: 100%;
    }
    .apply-button:hover {
        background-color: #45a049;
    }
    .edit-mode-container {
        background-color: #fff3cd;
        padding: 10px;
        border-radius: 5px;
        margin-bottom: 15px;
        border-left: 4px solid #ffc107;
    }
</style>
""", unsafe_allow_html=True)

# Dicionário de mapeamento de tipos de cobrança
def carregar_dicionario_mapeamento():
    return {
        # Encargo
        "Energia elet. adquirida 3": "Encargo",
        "energia elet adquirida 3": "Encargo",
        "energiaelet.adquirida3": "Encargo",
        "energia elet. adquirida3": "Encargo",
        "energia elet. adquirida 3* c/imposto": "Encargo",
        "energia elet. adquirida 3* c/lmposto": "Encargo",
        "energia elet. adquirida 3* c imp": "Encargo",
        "energia elet adquirida 3 c imposto": "Encargo",
        "Liminar Red Icms Fatura": "Encargo",
        "liminar red icms fatura": "Encargo",
        "liminarredicmsfatura": "Encargo",
        "liminar red. icms fatura": "Encargo",
        "Subtotal Faturamento": "Encargo",
        "Componente Encargo kWh HFP": "Encargo",
        "componente encargo kwh hfp": "Encargo",
        "Componente Encargo kWh HP": "Encargo",
        "componente encargo kwh hp": "Encargo",
        "Energia Terc Comercializad HFP": "Encargo",
        "energia terc comercializad hfp": "Encargo",
        "Energia Terc Comercializad HP": "Encargo",
        "energia terc comercializad hp": "Encargo",
        "Energia Isenta Comercial. HFP": "Encargo",
        "energia isenta comercial. hfp": "Encargo",
        "PIS/COFINS da subvenção/descon": "Encargo",
        "pis/cofins da subvenção/descon": "Encargo",
        "subtotal faturamento": "Encargo",
        "sub total faturamento": "Encargo",
        "subtot. faturamento": "Encargo",
        "ENERGIA ACL": "Encargo",
        "energia acl": "Encargo",
        "energiaacl": "Encargo",
        "TUSD ponta": "Encargo",
        "TUSD fora ponta": "Encargo",
        "tusd ponta": "Encargo",
        "tusd fora ponta": "Encargo",
        "Consumo Energia Elétrica HP": "Encargo",
        "consumo energia eletrica hp": "Encargo",
        "consumo energia elétrica hp": "Encargo",
        "DEMANDALEIESTADUAL16.886/18": "Encargo",
        
        # Desc. Encargo
        "DEDUCAO ENERGIA ACL": "Desc. Encargo",
        "DEDUCÃO ENERGIA ACL": "Desc. Encargo",
        "DED. ENERGIA ACL": "Desc. Encargo",
        "D. ENERGIA ACL": "Desc. Encargo",
        "Valor de Autoprodução HFP": "Desc. Encargo",
        "BENEFICIO TARIFARIO LIQUIDO": "Desc. Encargo",
        "BENEFÍCIO TARIFÁRIO LIQUIDO": "Desc. Encargo",
        "BENEFiCIO TARIFARIO LIQUIDO": "Desc. Encargo",
        "BENEFÍCIO TARIFÁRIO LÍQUIDO": "Desc. Encargo",
        "BENEFICIO TARIFÁRIO LÍQUIDO": "Desc. Encargo",
        "liminar icm5 tusd-consumo": "Desc. Encargo",
        "liminar icms tusd-consu mo": "Desc. Encargo",
        "liminar icms tusd-con5umo": "Desc. Encargo",
        "liminaricms tusd consumo": "Desc. Encargo",
        "liminar icms t usd-consumo": "Desc. Encargo",
        "liminar icms tu sd consumo": "Desc. Encargo",
        "liminar icms tusd-con sumo": "Desc. Encargo",
        "liminar icms t u s d-consumo": "Desc. Encargo",
        "liminar icms tusd-demanda": "Desc. Encargo",
        "liminar icms tusd demanda": "Desc. Encargo",
        "liminar icms tusddemanda": "Desc. Encargo",
        "liminar icms tus ddemanda": "Desc. Encargo",
        "liminar 1cms tusd-demanda": "Desc. Encargo",
        "liminar icms tuso-demanda": "Desc. Encargo",
        "liminar icm5 tusd-demanda": "Desc. Encargo",
        "liminar icms tusd demand a": "Desc. Encargo",
        "liminaricms tusd demanda": "Desc. Encargo",
        "liminar icms tu sd-demanda": "Desc. Encargo",
        "liminar icms t usd-demanda": "Desc. Encargo",
        "liminar icms tus-d-demanda": "Desc. Encargo",
        "liminar icms tusd d emanda": "Desc. Encargo",
        "liminar icms t u s d-demanda": "Desc. Encargo",
        "Liminar ICMS TUSD-Demanda": "Desc. Encargo",
        "Liminar ICMS TUSD Demanda": "Desc. Encargo",
        "Liminar ICMS TUSDDemanda": "Desc. Encargo",
        "Liminar ICMS TUS D Demanda": "Desc. Encargo",
        "Liminar ICMS TUS-D-Demanda": "Desc. Encargo",
        "Liminar ICMS TUSD Deman da": "Desc. Encargo",
        "Liminar 1CMS TUSD-Demanda": "Desc. Encargo",
        "Llminar ICMS TUSD-Demanda": "Desc. Encargo",
        "Liminar ICMS TUSO-Demanda": "Desc. Encargo",
        "Liminar ICMS TUS-Demanda": "Desc. Encargo",
        "Liminar ICMS TUSDDemand a": "Desc. Encargo",
        "Liminar ICMS TUSD-Consumo": "Desc. Encargo",
        "Liminar ICMS TUSD Consumo": "Desc. Encargo",
        "DEDUCAO ENERGIAACL": "Desc. Encargo",
        "DEDUÇÃO ENERGIAACL": "Desc. Encargo",
        "DED. ENERGIAACL": "Desc. Encargo",
        "deducao energiaacl": "Desc. Encargo",
        "deducaoenergiaacl":  "Desc. Encargo",
        "deducao energia acl": "Desc. Encargo",
        "DEDUÇÃO ENERGIA ACL": "Desc. Encargo",
        "Liminar ICMS TUSDConsumo": "Desc. Encargo",
        "Liminar ICMS TUS D Consumo": "Desc. Encargo",
        "Liminar ICMS TUSD-Consu mo": "Desc. Encargo",
        "Liminar ICMS TUSO-Consumo": "Desc. Encargo",
        "Liminar 1CMS TUSD-Consumo": "Desc. Encargo",
        "Llminar ICMS TUSD-Consumo": "Desc. Encargo",
        "Liminar ICMS TUSD-Consuno": "Desc. Encargo",
        "Liminar ICMS TUS DConsumo": "Desc. Encargo",
        "Valor de Autoproducao FP": "Desc. Encargo",
        "Valor de Autoprodução HP": "Desc. Encargo",
        "BENEFiCIO TARIFARIO LIQUIDO": "Desc. Encargo",
        "BENEFÍCIO TARIFÁRIO LIQUIDO": "Desc. Encargo",
        "BENEFÍCIO TARIFARIO LIQUIDO": "Desc. Encargo",
        "BENEFICIO TARIFARIO LIQUIDO": "Desc. Encargo",
        
        # Demanda
        "demanda ponta c/desconto": "Demanda",
        "demanda ponta cidesconto": "Demanda",
        "demanda ponta cidescon": "Demanda",
        "demanda ponta c/desc0nt0": "Demanda",
        "DEMANDA FORA PONTA C/DESC": "Demanda",
        "DEMANDA FORA PONTA C/DESC0": "Demanda",
        "DEMANDA FORA PONTA C/DESCONTO": "Demanda",
        "DEMANDA FORA PONTA C/DESC0NT0": "Demanda",
        "DEMANDA FORA PONTA CIDESCONTO": "Demanda",
        "DEMANDA FORA PONTA CIDESC0NTO": "Demanda",
        "Componente Fio kW HFP": "Demanda",
        "componente fio kw hfp": "Demanda",
        "Componente Fio kW HP": "Demanda",
        "componente fio kw hp": "Demanda",
        "demanda p0nta c/desconto": "Demanda",
        "demanda ponta c / desconto": "Demanda",
        "ULTRAPASSAGEM PONTA/TUSD": "Demanda",
        "ULTRAPASSAGEM PONTA TUSD": "Demanda",
        "ULTRAPASSAGEM PONTA TUSD-": "Demanda",
        "ULTRAPASSAGEM PONTA/TUSD ": "Demanda",
        "ULTRAPASSAGEM PONTA/T USD": "Demanda",
        "ULTRAPASSAGEM PONTA/TUSD-": "Demanda",
        "ULTRAPASSAGEM PONTA/TUSD  ": "Demanda",
        "ULTRAPASSAGEM PONTA-TUSD": "Demanda",
        "ULTRA PASSAGEM PONTA/TUSD": "Demanda",
        "ULTRA PASSAGEM PONTA TUSD": "Demanda",
        "ULTRAPASSAGEM PONTA TUSD ": "Demanda",
        "ULTRAPASSAGEM FORAPONTA/TUSD": "Demanda",
        "ULTRAPASSAGEM FORA PONTA TUSD": "Demanda",
        "ULTRAPASSAGEM FORAPONTA TUSD": "Demanda",
        "ULTRAPASSAGEM FORA PONTA/TUSD": "Demanda",
        "ULTRAPASSAGEM FORA PONTA TUSD ": "Demanda",
        "UFDR PONTA/TUSD": "Demanda",
        "UFDRPONTA/TUSD": "Demanda",
        "UFDR PONTA TUSD": "Demanda",
        "UFDRPONTATUSD": "Demanda",
        "UFDR P0NTA/TUSD": "Demanda",
        "UFDR PONTA/T USD": "Demanda",
        "UFDR P0NTA TUSD": "Demanda",
        "UFDR PONTA-TUSD": "Demanda",
        "UFDR PONTA/TUS D": "Demanda",
        "UFDRFORA PONTA/TUSD": "Demanda",
        "UFDR FORA PONTA/TUSD": "Demanda",
        "UFDR FORAPONTA/TUSD": "Demanda",
        "UFDR FORA PONTA TUSD": "Demanda",
        "UFDRFORAPONTA TUSD": "Demanda",
        "UFDRFORA PONTA TUSD": "Demanda",
        "UFDR FORA P0NTA/TUSD": "Demanda",
        "UFDRFORA PONTA-TUSD": "Demanda",
        "UFDR FORA PONTA/TUS D": "Demanda",
        "ULTRAPASSAGEM FORA PONTA T USD": "Demanda",
        "ULTRAPASSAGEM FORA PONTA TUSD-": "Demanda",
        "ULTRAPASSAGEM FORA PONTA/TUSD ": "Demanda",
        "ULTRAPASSAGEMFORAPONTA TUSD": "Demanda",
        "ULTRAPASSAGEMF.PONTA/TUSD": "Demanda",
        "ULTRAPASSAGEM FORA PONTA/TUSD-": "Demanda",
        "ULTRA PASSAGEM FORA PONTA/TUSD": "Demanda",
        "demanda pontac/desconto": "Demanda",
        "demanda ponta c/desconto": "Demanda",
        "demanda ponta cidesconto": "Demanda",
        "demanda ponta cidescon": "Demanda",
        "demanda ponta c/desc0nt0": "Demanda",
        "Componente Fio kW HFP": "Demanda",
        "Componente Fio kW HP": "Demanda",
        "Component Fio kW HFP": "Demanda",
        "Component Fio kW HP": "Demanda",
        "Componente Fio kWHP": "Demanda",
        "Componente Fio KW HP": "Demanda",
        "Componente Fio KW HFP": "Demanda",
        "Componete Fio kW HFP": "Demanda",
        "Comp Fio kW HFP": "Demanda",
        "Comp Fio kW HP": "Demanda",
        "Componente Fio kw hfp": "Demanda",
        "demanda p0nta c/desconto": "Demanda",
        "demanda ponta c / desconto": "Demanda",
        "demanda f0ra p0nta cidescont0": "Demanda",
        "demanda fora ponta cidesconto": "Demanda",
        "demanda fora p0nta cidesconto": "Demanda",
        "demanda fora ponta cidesc0nto": "Demanda",
        "demanda fora ponta cidescon": "Demanda",
        "demanda f ora ponta cidesconto": "Demanda",
        "demanda fora pontacidesconto": "Demanda",
        "DEMANDA PONTA C/DESCONTO": "Demanda",
        "DEMANDA PONTA C DESCONTO": "Demanda",
        "DEMANDA PONTA C/DESC0NT0": "Demanda",
        "DEMANDA P0NTA C/DESCONTO": "Demanda",
        "DEMANDA P0NTA C DESCONTO": "Demanda",
        "DEMANDA FORA PONTA CIDESCONTO": "Demanda",
        "DEMANDA F0RA P0NTA CIDESCONTO": "Demanda",
        "DEMANDA FORA PONTA C/DESC": "Demanda",
        "DEMANDA  FORA  PONTA  C/DESC": "Demanda",
        "DEMANDA FORA PONTA C/DESC0NT0": "Demanda",
        "Consumo Energia Eletrica HFP": "Demanda",
        "Consumo de Energia Eletrica HFP": "Demanda",
        "Consumo Energia Eletrica HFP/Unico": "Demanda",
        "Consumo de Energia Elétrica HFP": "Demanda",
        "Consumo Energia Elétric HFP": "Demanda",
        "Consumo Energia Eléctrica HFP": "Demanda",
        "Consumo Energia Eletrica kWh HFP": "Demanda",
        "Componente Fio HFP s/ ICMS": "Demanda",
        "Componente Fio HFP": "Demanda",
        "Componente Fio HP": "Demanda",
        "Componente Fio HFP s/ICMS": "Demanda",
        "Componente Fio kW HFP": "Demanda",
        "Componente Fio kW HP": "Demanda",
        "Component Fio kW HFP": "Demanda",
        "Component Fio kW HP": "Demanda",
        "Componente Fio kWHP": "Demanda",
        "Componente Fio KW HP": "Demanda",
        "Componente Fio KW HFP": "Demanda",
        "Componete Fio kW HFP": "Demanda",
        "Comp Fio kW HFP": "Demanda",
        "Comp Fio kW HP": "Demanda",
        "Componente Fio kw hfp": "Demanda",
        "Demanda Ativa kW HFP/Unico": "Demanda",
        "Demanda Ativa kW HFP/unico": "Demanda",
        "Demanda Ativa kW HFP/ünico": "Demanda",
        "Demanda Ativa kW HFP/Un1co": "Demanda",
        "Demanda Ativa kW HFP/Unic0": "Demanda",
        "Demanda AtivakW HFP/Unico": "Demanda",
        "Demanda Ativa kWHFP/Unico": "Demanda",
        "Componente Fio kW HFP": "Demanda",
        "Componente Fio kW HP": "Demanda",
        "Component Fio kW HFP": "Demanda",
        "Component Fio kW HP": "Demanda",
        "Componente Fio kWHP": "Demanda",
        "Componente Fio KW HP": "Demanda",
        "Componente Fio KW HFP": "Demanda",
        "Componete Fio kW HFP": "Demanda",
        "Comp Fio kW HFP": "Demanda",
        "Comp Fio kW HP": "Demanda",
        "Componente Fio kw hfp": "Demanda",
        "Componente Fio HFP": "Demanda",
        "Componente Fio HP": "Demanda",
        "Componente Fio kW HP": "Demanda",
        "Componente Fio HFP s/ICMS": "Demanda",
        "Componente Fio HP s/ICMS": "Demanda",
        "Componente Fio HP s/ ICMS": "Demanda",
        "Componente Fio HFP s/ ICMS": "Demanda",
        "DEMANDA PONTA C/DESCONTO": "Demanda",
        "DEMANDA FORA PONTA CIDESCONTO": "Demanda",
        "DEMANDA UNICA C/DESCONTO": "Demanda",
        "Demanda Distribuicao F. Ponta": "Demanda",
        "Demanda Distribuicao Ponta": "Demanda",
        "Dernanda Distribuicao Ponta": "Demanda",
        "DIF.FATUR.TUSD-HOMOLOG.CCEE": "Demanda",
        "ULTRAPASSAGEM FORAPONTA/TUSD": "Demanda",
        
        # Desc. Demanda
        "DIF.FATUR.TUSD-HOMOLOG.CCEE": "Desc. Demanda",
        "DIF.FATUR.TUSD-HOMOLOG.C0EE": "Desc. Demanda",
        "DIF.FATUR.TUSD-HOMOLOG.0CEE": "Desc. Demanda",
        "DIF.FATUR.TUSD HOMOLOG CCEE": "Desc. Demanda",
        "DIF.FATUR TUSD-HOMOLOG CCEE": "Desc. Demanda",
        "DIF.FATUR.T0SD-HOMOLOG.CCEE": "Desc. Demanda",
        "D1F.FATUR.TUSD-H0MOLOG.CCEE": "Desc. Demanda",
        "DIF.FATUR.TUSD HOMOLOG.CCEE": "Desc. Demanda",
        "DIF FATUR.TUSD-HOMOLOG.CCEE": "Desc. Demanda",
        "DIF.FATUR.TUSD-HOMOLOG.C0E": "Desc. Demanda",
        "DIF.FATUR.TUSD-HOMOLOG.CEE": "Desc. Demanda",
        "DIF.FATUR.TUSD-HOMOLOG.CCE": "Desc. Demanda",
        "ajuste de desconto cfio hfp": "Desc. Demanda",
        "Ajuste de Desconto C. Fio HFP": "Desc. Demanda",
        "ajuste de desconto c. fio hfp": "Desc. Demanda",
        "Ajuste de Desconto C. Fio HP": "Desc. Demanda",
        "ajuste de desconto c. fio hp": "Desc. Demanda",
        "Desconto Comp. Fio HFP": "Desc. Demanda",
        "desconto comp. fio hfp": "Desc. Demanda",
        "Desconto Comp. Fio HP": "Desc. Demanda",
        "desconto comp. fio hp": "Desc. Demanda",
        
        # Reativa
        "ENERGIA REATIVA KVARH HFP": "Reativa",
        "ENERGIA REATIVA KVARH HP": "Reativa",
        "ENERGIA REATIVA KVARH": "Reativa",
        "ENERGIA REATIVA KVARH HFP/UNICO": "Reativa",
        "ENERGIA REATIVA KVARH HP/UNICO": "Reativa",
        "ENERGIA REATIVA KVARH/UNICO": "Reativa",
        "ENERGIA REATIVA KVARH HFP/ÚNICO": "Reativa",
        "ENERGIA REATIVA KVARH HP/ÚNICO": "Reativa",
        "ENERGIA REATIVA KVARH/ÚNICO": "Reativa",
        "ENERGIA REATIVA KVARH HFP/UNlCO": "Reativa",
        "ENERGIA REATIVA KVARH HP/UNlCO": "Reativa",
        "ENERGIA REATIVA KVARH/UNlCO": "Reativa",
        "ENERGIA REATIVA KVARH HFP/UN1CO": "Reativa",
        "ENERGIA REATIVA KVARH HP/UN1CO": "Reativa",
        "ENERGIA REATIVA KVARH/UN1CO": "Reativa",
        "ENERGIA REATIVA KVARH HFP/UNIC0": "Reativa",
        "ENERGIA REATIVA KVARH HP/UNIC0": "Reativa",
        "ENERGIA REATIVA KVARH/UNIC0": "Reativa",
        "ENERGIA REATIVA KVARH HFP/UNlC0": "Reativa",
        "ENERGIA REATIVA KVARH HP/UNlC0": "Reativa",
        "ENERGIA REATIVA KVARH/UNlC0": "Reativa",
        "ENERGIA REATIVA KVARH HFP/UN1C0": "Reativa",
        "ENERGIA REATIVA KVARH HP/UN1C0": "Reativa",
        "ENERGIA REATIVA KVARH/UN1C0": "Reativa",
        "ENERGIA REATIVA KVARH HFP/UNICO": "Reativa",
        "ENERGIA REATIVA KVARH HP/UNICO": "Reativa",
        "ENERGIA REATIVA KVARH/UNICO": "Reativa",
        "ENERGIA REATIVA KVARH HFP/ÚNICO": "Reativa",
        "ENERGIA REATIVA KVARH HP/ÚNICO": "Reativa",
        "ENERGIA REATIVA KVARH/ÚNICO": "Reativa",
        "ENERGIA REATIVA KVARH HFP/UNlCO": "Reativa",
        "ENERGIA REATIVA KVARH HP/UNlCO": "Reativa",
        "ENERGIA REATIVA KVARH/UNlCO": "Reativa",
        "ENERGIA REATIVA KVARH HFP/UN1CO": "Reativa",
        "ENERGIA REATIVA KVARH HP/UN1CO": "Reativa",
        "ENERGIA REATIVA KVARH/UN1CO": "Reativa",
        "ENERGIA REATIVA KVARH HFP/UNIC0": "Reativa",
        "ENERGIA REATIVA KVARH HP/UNIC0": "Reativa",
        "ENERGIA REATIVA KVARH/UNIC0": "Reativa",
        "ENERGIA REATIVA KVARH HFP/UNlC0": "Reativa",
        "ENERGIA REATIVA KVARH HP/UNlC0": "Reativa",
        "ENERGIA REATIVA KVARH/UNlC0": "Reativa",
        "ENERGIA REATIVA KVARH HFP/UN1C0": "Reativa",
        "ENERGIA REATIVA KVARH HP/UN1C0": "Reativa",
        "ENERGIA REATIVA KVARH/UN1C0": "Reativa",
        "ENERGIA REATIVA KWH HFP": "Reativa",
        "ENERGIA REATIVA KWH H P": "Reativa",
        "ENERGIA REATIVA KWH HFPUNICO": "Reativa",
        "ENERGIA REATIVA KW HFP": "Reativa",
        "ENERGIA REATIVA KWH": "Reativa",
        "ENERGIAREATIVA KWH": "Reativa",
        "ENERGIA REATIVA HP ": "Reativa",
        "UFER PONTA TEE": "Reativa",
        "UFER FORA PONTA TE": "Reativa",
        "UFER FORA P0NTA TE": "Reativa",
        "UFER FORA PONTA-TE": "Reativa",
        "UFER FORA PONTA.TE": "Reativa",
        "UFER-FORA PONTA TE": "Reativa",
        "UFER PONTA TE (NÃO FATURADO)": "Reativa",
        
        # Contb. Publica
        "CIP-ILUM PUB PREF MUNICIPAL": "Contb. Publica",
        "CIP - ILUM PUB PREF MUNICIPAL": "Contb. Publica",
        "COSIP.SAOPAULO-MUNICIPAL": "Contb. Publica",
        "CIP-BARUERI-MUNICIPAL": "Contb. Publica",
        "CIP-CAJAMAR-MUNICIPAL": "Contb. Publica",
        "CIP-ILUM PUB PREF MUNICIPALCONTRIB PUB": "Contb. Publica",
        "CIP-SAOPAULO-MUNICIPAL": "Contb. Publica",
        "CIP-ILUM PUB PREF MUNICIPAL-SAO PAULO": "Contb. Publica",
        "CIP_ILUM_PUB_MUNICIPAL": "Contb. Publica",
        "Contrib llum Publica Municipal": "Contb. Publica",
        "Contrib Ilum Publica Municipal": "Contb. Publica",
        "Contrib Ilum Publica Municipal": "Contb. Publica",
        "Contrib IlumPublica Municipal": "Contb. Publica",
        "Contrib IlumPublicaMunicipal": "Contb. Publica",
        "Contrib llumPublica Municipal": "Contb. Publica",
        "Contrib llumPublicaMunicipal": "Contb. Publica",
        "Contrib Ilum Publica Municipal": "Contb. Publica",
        "contrib ilum publica municipal": "Contb. Publica",
        "Contrib Ilum Publica Municipal": "Contb. Publica",
        "contrib ilum publica municipal": "Contb. Publica",
        "CIPILUM PUB PREF MUNICIPAL": "Contb. Publica",
        "CIP ILUM PUB PREF MUNICIPAL": "Contb. Publica",
        "C1P - ILUM PUB PREF MUNICIPAL": "Contb. Publica",
        "C1P-ILUM PUB PREF MUNICIPAL": "Contb. Publica",
        "CIP ILUM PUBL PREF MUNICIPAL": "Contb. Publica",
        "COSIP SAOPAULO MUNICIPAL": "Contb. Publica",
        "COSIP SAO PAULO MUNICIPAL": "Contb. Publica",
        "C0SIP.SAOPAULO-MUNICIPAL": "Contb. Publica",
        "CIP-BARUERI MUNICIPAL": "Contb. Publica",
        "CIP CAJAMAR MUNICIPAL": "Contb. Publica",
        "CIP ILUM PUB MUNICIPAL": "Contb. Pblica",
        "CIP-ILUM PUB PREF MUNICIPAL CONTRIB PUBLICA": "Contb. Publica",
        "CIP ILUM PUBLICA MUNICIPAL": "Contb. Publica",
        "CIP-ILUM PUBLICA MUNICIPAL": "Contb. Publica",
        "Contrib llum Publica Municpal": "Contb. Publica",
        "Contrib llum Pubica Municipal": "Contb. Publica",
        "Contrib Ilum Publica Municpal": "Contb. Publica",
        "Contribuçao Ilum Publica": "Contb. Publica",
        "Contribuiçao Ilum Publica": "Contb. Publica",
        "Contribuiçâo llum Publica Municipal": "Contb. Publica",
        "Cont llum Publica": "Contb. Publica",
        "Contrib IlumPublicaMunicipal": "Contb. Publica",
        "Contr llumPublicaMunicipal": "Contb. Publica",
        "Contribuicao Ilum Publica": "Contb. Publica",
        "Contribuicao IlumPublica": "Contb. Publica",
        "Contribuição Ilum Publica Municipal": "Contb. Publica",
        "Cont Ilum Publica": "Contb. Publica",
        "C0ntrib Ilum Publica": "Contb. Publica",
        "Contrib llum Püblica Municipal": "Contb. Publica",
        "Contribuição Iluminação Publica": "Contb. Publica",
        "Cont llum Publica": "Contb. Publica",
        "Contribuição Publica": "Contb. Publica",
        "Contrib llum Publica": "Contb. Publica",
        "Contrib IlumMunicipal": "Contb. Publica",
        "Contrib llum Püblica Municipal": "Contb. Publica",
        "Contrib Ilum Publica Municipal": "Contb. Publica",
        "Contrib Ilum Publica Municipal": "Contb. Publica",
        "Contrib IlumPublica Municipal": "Contb. Publica",
        "Contrib llumPublica Municipal": "Contb. Publica",
        "Contrib IlumPüblica Municipal": "Contb. Publica",
        "Contrib llum Publica Municipal": "Contb. Publica",
        "Contrib llum Püblca Municipal": "Contb. Publica",
        "Contrib llum Publica Municpal": "Contb. Publica",
        "Contrib llum Püblca Municipal": "Contb. Publica",
        "Contrib llum Publica Municpal": "Contb. Publica",
        "Contrib llum Püblica Municipal": "Contb. Publica",
        "Contrib llum Publica Municpal": "Contb. Publica",
        "Contrib Ilum Publica Municipal": "Contb. Publica",
        "Contrib Ilum Publica Municipal": "Contb. Publica",
        "Contrib IlumPublica Municipal": "Contb. Publica",
        "Contrib llumPublica Municipal": "Contb. Publica",
        "Contrib IlumPüblica Municipal": "Contb. Publica",
        "Contrib llum Publica Municipal": "Contb. Publica",
        "Contrib llum Püblca Municipal": "Contb. Publica",
    }

# Lista de senhas conhecidas para tentar desbloquear PDFs protegidos
senhas_tentativas = ["33042", "56993", "60869", "08902", "3304", "5699", "6086", "0890"]

# Função para normalizar texto (remover acentos, converter para minúsculas)
def normalizar_texto(texto):
    if not texto:
        return ""
    texto = unicodedata.normalize('NFKD', texto).encode('ASCII', 'ignore').decode('ASCII')
    return texto.lower().strip()

# Função para tentar desbloquear um PDF protegido
def desbloquear_pdf(pdf_bytes):
    # Primeiro, verificar se o PDF está protegido
    try:
        with PyPDF2.PdfReader(io.BytesIO(pdf_bytes)) as reader:
            # Se conseguir ler a primeira página, não está protegido
            if len(reader.pages) > 0:
                reader.pages[0].extract_text()
                return pdf_bytes  # Retorna o PDF original se não estiver protegido
    except:
        # PDF está protegido, tentar desbloquear
        for senha in senhas_tentativas:
            try:
                with pikepdf.open(io.BytesIO(pdf_bytes), password=senha) as pdf:
                    output = io.BytesIO()
                    pdf.save(output)
                    return output.getvalue()
            except:
                continue
    
    # Se chegou aqui, não conseguiu desbloquear
    return None

# Função para extrair texto de uma área específica do PDF
def extrair_texto_area(pdf_bytes, pagina, x1, y1, x2, y2):
    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            if pagina < len(pdf.pages):
                page = pdf.pages[pagina]
                # Ajustar coordenadas para o sistema do pdfplumber (0,0 é canto inferior esquerdo)
                height = page.height
                
                # Validar coordenadas para garantir que estejam dentro dos limites da página
                x1 = max(0, min(x1, page.width))
                x2 = max(0, min(x2, page.width))
                y1 = max(0, min(y1, height))
                y2 = max(0, min(y2, height))
                
                crop_box = (x1, height - y2, x2, height - y1)
                cropped_page = page.crop(crop_box)
                return cropped_page.extract_text() or ""
            return ""
    except Exception as e:
        st.error(f"Erro ao extrair texto da área: {str(e)}")
        return ""

# Função para extrair tabela de uma área específica do PDF
def extrair_tabela_area(pdf_bytes, pagina, x1, y1, x2, y2):
    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            if pagina < len(pdf.pages):
                page = pdf.pages[pagina]
                # Ajustar coordenadas para o sistema do pdfplumber (0,0 é canto inferior esquerdo)
                height = page.height
                
                # Validar coordenadas para garantir que estejam dentro dos limites da página
                x1 = max(0, min(x1, page.width))
                x2 = max(0, min(x2, page.width))
                y1 = max(0, min(y1, height))
                y2 = max(0, min(y2, height))
                
                crop_box = (x1, height - y2, x2, height - y1)
                cropped_page = page.crop(crop_box)
                
                # Extrair texto para enviar à API da OpenAI
                texto = cropped_page.extract_text() or ""
                return texto
            return ""
    except Exception as e:
        st.error(f"Erro ao extrair tabela da área: {str(e)}")
        return ""

# Função para identificar a distribuidora pelo cabeçalho do PDF
def identificar_distribuidora(pdf_bytes):
    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            if len(pdf.pages) > 0:
                page = pdf.pages[0]
                texto = page.extract_text() or ""
                
                # Padrões de identificação de distribuidoras
                padroes = {
                    "ENEL": ["ENEL", "AMPLA", "COELCE", "ELETROPAULO"],
                    "LIGHT": ["LIGHT"],
                    "CEMIG": ["CEMIG"],
                    "CPFL": ["CPFL", "COMPANHIA PAULISTA DE FORÇA E LUZ"],
                    "ENERGISA": ["ENERGISA"],
                    "EDP": ["EDP", "ESPIRITO SANTO", "SÃO PAULO"],
                    "EQUATORIAL": ["EQUATORIAL", "CEMAR", "CELPA"],
                    "COPEL": ["COPEL"],
                    "CELESC": ["CELESC"],
                    "NEOENERGIA": ["NEOENERGIA", "COELBA", "COSERN", "CELPE", "ELEKTRO"]
                }
                
                for distribuidora, termos in padroes.items():
                    for termo in termos:
                        if termo in texto.upper():
                            return distribuidora
                
                # Se não encontrar, tentar identificar pelo nome do arquivo
                return "Não identificada"
    except Exception as e:
        st.error(f"Erro ao identificar distribuidora: {str(e)}")
        return "Não identificada"

# Função para processar o texto extraído com a API da OpenAI
def processar_com_openai(api_key, texto, modelo="gpt-3.5-turbo"):
    if not texto.strip():
        return None, 0, 0
    
    try:
        cliente = OpenAI(api_key=api_key)
        
        prompt = f"""
        Extraia a tabela de itens de faturamento do seguinte texto de uma fatura de energia elétrica.
        Retorne apenas os dados em formato JSON com as seguintes colunas: Descrição, Quantidade, Tarifa, Valor.
        Se houver outras colunas relevantes, inclua-as também.
        
        Texto da fatura:
        {texto}
        
        Formato de resposta:
        {{
            "itens": [
                {{
                    "Descrição": "texto",
                    "Quantidade": "valor",
                    "Tarifa": "valor",
                    "Valor": "valor"
                }}
            ]
        }}
        """
        
        response = cliente.chat.completions.create(
            model=modelo,
            messages=[
                {"role": "system", "content": "Você é um assistente especializado em extrair dados estruturados de faturas de energia elétrica."},
                {"role": "user", "content": prompt}
            ],
            temperature=0.1
        )
        
        # Calcular custo
        input_tokens = response.usage.prompt_tokens
        output_tokens = response.usage.completion_tokens
        
        # Preços aproximados (podem variar)
        if modelo == "gpt-3.5-turbo":
            input_cost = input_tokens * 0.0000015  # $0.0015 por 1000 tokens
            output_cost = output_tokens * 0.000002  # $0.002 por 1000 tokens
        else:  # gpt-4
            input_cost = input_tokens * 0.00003  # $0.03 por 1000 tokens
            output_cost = output_tokens * 0.00006  # $0.06 por 1000 tokens
        
        total_cost = input_cost + output_cost
        
        # Extrair o conteúdo JSON da resposta
        content = response.choices[0].message.content
        
        # Tentar encontrar o JSON na resposta
        try:
            # Procurar por padrões de JSON na resposta
            json_match = re.search(r'({.*})', content, re.DOTALL)
            if json_match:
                json_str = json_match.group(1)
                dados = json.loads(json_str)
                return dados, input_tokens + output_tokens, total_cost
            else:
                return None, input_tokens + output_tokens, total_cost
        except Exception as e:
            st.error(f"Erro ao processar resposta da API: {str(e)}")
            return None, input_tokens + output_tokens, total_cost
            
    except Exception as e:
        st.error(f"Erro ao chamar a API da OpenAI: {str(e)}")
        return None, 0, 0

# Função para classificar itens de acordo com o dicionário de mapeamento
def classificar_itens(dados, dicionario_mapeamento):
    if not dados or "itens" not in dados:
        return {}
    
    resultado = {
        "Encargo": 0.0,
        "Desc. Encargo": 0.0,
        "Demanda": 0.0,
        "Desc. Demanda": 0.0,
        "Reativa": 0.0,
        "Contb. Publica": 0.0,
        "Outros": 0.0
    }
    
    for item in dados["itens"]:
        if "Descrição" in item and "Valor" in item:
            descricao = normalizar_texto(item["Descrição"])
            valor_str = item["Valor"].replace("R$", "").replace(".", "").replace(",", ".").strip()
            
            try:
                valor = float(valor_str)
            except:
                valor = 0.0
            
            # Classificar o item
            classificado = False
            for padrao, categoria in dicionario_mapeamento.items():
                if padrao.lower() in descricao:
                    resultado[categoria] += valor
                    classificado = True
                    break
            
            # Se não foi classificado, vai para "Outros"
            if not classificado:
                resultado["Outros"] += valor
    
    return resultado

# Função para converter o DataFrame para Excel
def to_excel(df):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Dados')
        workbook = writer.book
        worksheet = writer.sheets['Dados']
        
        # Formatar cabeçalhos
        for col_num, column_title in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color='E6E6E6', end_color='E6E6E6', fill_type='solid')
        
        # Ajustar largura das colunas
        for i, col in enumerate(df.columns):
            column_width = max(df[col].astype(str).map(len).max(), len(col)) + 2
            worksheet.column_dimensions[openpyxl.utils.get_column_letter(i+1)].width = column_width
    
    processed_data = output.getvalue()
    return processed_data

# Função para converter PDF para imagem
def pdf_para_imagem(pdf_bytes, pagina=0, dpi=200):
    try:
        images = convert_from_bytes(pdf_bytes, dpi=dpi, first_page=pagina+1, last_page=pagina+1)
        if images:
            return images[0]
        return None
    except Exception as e:
        st.error(f"Erro ao converter PDF para imagem: {str(e)}")
        return None

# Função para desenhar retângulos de seleção na imagem
def desenhar_selecao(img, ref_coords=None, tab_coords=None):
    # Criar uma cópia da imagem para não modificar a original
    img_copy = img.copy()
    draw = ImageDraw.Draw(img_copy, 'RGBA')
    
    # Desenhar retângulo de referência
    if ref_coords and len(ref_coords) == 4:
        x1, y1, x2, y2 = ref_coords
        draw.rectangle([x1, y1, x2, y2], outline="blue", width=3, fill=(0, 0, 255, 50))
    
    # Desenhar retângulo de tabela
    if tab_coords and len(tab_coords) == 4:
        x1, y1, x2, y2 = tab_coords
        draw.rectangle([x1, y1, x2, y2], outline="green", width=3, fill=(0, 255, 0, 50))
    
    return img_copy

# Função para salvar padrões de seleção
def salvar_padrao_selecao(distribuidora, ref_coords, tab_coords):
    padroes = {}
    
    # Carregar padrões existentes, se houver
    if 'padroes_selecao' in st.session_state:
        padroes = st.session_state.padroes_selecao
    
    # Adicionar ou atualizar padrão para esta distribuidora
    padroes[distribuidora] = {
        "referencia": ref_coords,
        "tabela": tab_coords
    }
    
    # Salvar na sessão
    st.session_state.padroes_selecao = padroes

# Função para carregar padrões de seleção
def carregar_padrao_selecao(distribuidora):
    if 'padroes_selecao' in st.session_state and distribuidora in st.session_state.padroes_selecao:
        return st.session_state.padroes_selecao[distribuidora]
    return None

# Inicialização da sessão
if 'resultados' not in st.session_state:
    st.session_state.resultados = []

if 'padroes_selecao' not in st.session_state:
    st.session_state.padroes_selecao = {}

if 'total_tokens' not in st.session_state:
    st.session_state.total_tokens = 0

if 'total_custo' not in st.session_state:
    st.session_state.total_custo = 0.0

if 'pdf_atual' not in st.session_state:
    st.session_state.pdf_atual = None

if 'edit_mode' not in st.session_state:
    st.session_state.edit_mode = False

# Interface principal
st.title("📊 Leitor de Faturas de Energia")

# Sidebar para configurações
with st.sidebar:
    st.header("Configurações")
    api_key = st.text_input("Chave da API OpenAI", type="password")
    modelo = st.selectbox("Modelo OpenAI", ["gpt-3.5-turbo", "gpt-4"], index=0)
    
    st.header("Instruções")
    st.markdown("""
    1. Insira sua chave da API OpenAI
    2. Faça upload das faturas em PDF
    3. Para cada fatura:
       - Ajuste as áreas de seleção (azul e verde)
       - Clique em "Aplicar seleção"
       - Clique em "Processar esta fatura"
    4. Visualize e exporte os resultados
    """)
    
    st.header("Sobre")
    st.markdown("""
    Esta aplicação extrai dados de faturas de energia elétrica usando OCR e processamento de linguagem natural.
    
    Desenvolvido para otimizar o processamento de faturas com estrutura variável.
    """)

# Área principal
st.header("Upload de Faturas")

# Upload de arquivos
uploaded_files = st.file_uploader("Selecione os arquivos PDF", type="pdf", accept_multiple_files=True)

if uploaded_files:
    st.success(f"{len(uploaded_files)} arquivos carregados com sucesso!")
    
    # Botão para processar faturas
    if st.button("Processar Faturas", key="processar_faturas"):
        if not api_key:
            st.error("Por favor, insira sua chave da API OpenAI para continuar.")
        else:
            # Resetar resultados anteriores
            st.session_state.resultados = []
            st.session_state.total_tokens = 0
            st.session_state.total_custo = 0.0
            
            progress_bar = st.progress(0)
            status_text = st.empty()
            
            # Carregar dicionário de mapeamento
            dicionario = carregar_dicionario_mapeamento()
            
            # Processar cada arquivo
            for i, uploaded_file in enumerate(uploaded_files):
                status_text.text(f"Processando {uploaded_file.name}... ({i+1}/{len(uploaded_files)})")
                
                # Ler o arquivo PDF
                pdf_bytes = uploaded_file.read()
                
                # Tentar desbloquear se estiver protegido
                pdf_desbloqueado = desbloquear_pdf(pdf_bytes)
                if pdf_desbloqueado is None:
                    st.error(f"Não foi possível desbloquear o arquivo {uploaded_file.name}. Verifique se a senha está na lista de senhas conhecidas.")
                    continue
                
                # Identificar a distribuidora
                distribuidora = identificar_distribuidora(pdf_desbloqueado)
                
                # Verificar se há padrão de seleção salvo para esta distribuidora
                padrao = carregar_padrao_selecao(distribuidora)
                
                if padrao:
                    # Usar padrão salvo
                    ref_coords = padrao["referencia"]
                    tab_coords = padrao["tabela"]
                    
                    # Extrair texto da área de referência
                    texto_ref = extrair_texto_area(pdf_desbloqueado, 0, *ref_coords)
                    
                    # Extrair texto da área da tabela
                    texto_tab = extrair_tabela_area(pdf_desbloqueado, 0, *tab_coords)
                    
                    # Processar com OpenAI
                    dados, tokens, custo = processar_com_openai(api_key, texto_tab, modelo)
                    
                    # Atualizar contadores
                    st.session_state.total_tokens += tokens
                    st.session_state.total_custo += custo
                    
                    # Classificar itens
                    if dados:
                        classificacao = classificar_itens(dados, dicionario)
                        
                        # Extrair informações de referência e instalação
                        padrao_instalacao = r'Instalação:\s*(\d+)'
                        padrao_referencia = r'Referente a\s*([A-Za-z]+/\d{4})'
                        
                        instalacao = re.search(padrao_instalacao, texto_ref)
                        instalacao = instalacao.group(1) if instalacao else "N/A"
                        
                        referencia = re.search(padrao_referencia, texto_ref)
                        referencia = referencia.group(1) if referencia else "N/A"
                        
                        # Adicionar aos resultados
                        st.session_state.resultados.append({
                            "Arquivo": uploaded_file.name,
                            "Distribuidora": distribuidora,
                            "Instalação": instalacao,
                            "Referência": referencia,
                            "Dados": dados,
                            "Classificação": classificacao,
                            "Tokens": tokens,
                            "Custo": custo
                        })
                    else:
                        st.warning(f"Não foi possível extrair dados estruturados de {uploaded_file.name}")
                else:
                    # Exibir PDF para seleção manual
                    st.subheader(f"Selecione áreas em: {uploaded_file.name}")
                    
                    # Salvar PDF atual na sessão
                    st.session_state.pdf_atual = {
                        "nome": uploaded_file.name,
                        "bytes": pdf_desbloqueado,
                        "distribuidora": distribuidora
                    }
                    
                    # Converter PDF para imagem
                    img = pdf_para_imagem(pdf_desbloqueado)
                    
                    if img:
                        # Inicializar coordenadas se não existirem
                        if f"ref_coords_{i}" not in st.session_state:
                            st.session_state[f"ref_coords_{i}"] = [
                                int(img.width * 0.1),
                                int(img.height * 0.1),
                                int(img.width * 0.9),
                                int(img.height * 0.2)
                            ]
                        
                        if f"tab_coords_{i}" not in st.session_state:
                            st.session_state[f"tab_coords_{i}"] = [
                                int(img.width * 0.1),
                                int(img.height * 0.3),
                                int(img.width * 0.9),
                                int(img.height * 0.7)
                            ]
                        
                        # Inicializar modo de edição
                        if f"edit_mode_{i}" not in st.session_state:
                            st.session_state[f"edit_mode_{i}"] = False
                        
                        # Inicializar valores temporários
                        if f"temp_ref_coords_{i}" not in st.session_state:
                            st.session_state[f"temp_ref_coords_{i}"] = st.session_state[f"ref_coords_{i}"].copy()
                        
                        if f"temp_tab_coords_{i}" not in st.session_state:
                            st.session_state[f"temp_tab_coords_{i}"] = st.session_state[f"tab_coords_{i}"].copy()
                        
                        # Instruções para seleção
                        st.markdown("""
                        <div class="selection-instructions">
                            <div class="step-container">
                                <span class="step-indicator">1</span>
                                <span class="step-text">Clique em "Editar seleção" para ajustar as áreas</span>
                            </div>
                            <div class="step-container">
                                <span class="step-indicator">2</span>
                                <span class="step-text">Ajuste os sliders para selecionar as áreas (azul e verde)</span>
                            </div>
                            <div class="step-container">
                                <span class="step-indicator">3</span>
                                <span class="step-indicator">3</span>
                                <span class="step-text">Clique em "Aplicar seleção" para confirmar</span>
                            </div>
                            <div class="step-container">
                                <span class="step-indicator">4</span>
                                <span class="step-text">Clique em "Processar esta fatura" quando terminar</span>
                            </div>
                        </div>
                        """, unsafe_allow_html=True)
                        
                        # Criar contêiner para a imagem e seleção
                        col1, col2 = st.columns([2, 1])
                        
                        with col1:
                            # Desenhar retângulos de seleção na imagem
                            if st.session_state[f"edit_mode_{i}"]:
                                # No modo de edição, usar valores temporários
                                img_com_selecao = desenhar_selecao(
                                    img, 
                                    st.session_state[f"temp_ref_coords_{i}"],
                                    st.session_state[f"temp_tab_coords_{i}"]
                                )
                            else:
                                # Fora do modo de edição, usar valores confirmados
                                img_com_selecao = desenhar_selecao(
                                    img, 
                                    st.session_state[f"ref_coords_{i}"],
                                    st.session_state[f"tab_coords_{i}"]
                                )
                            
                            # Exibir imagem com seleções
                            st.image(img_com_selecao, caption=f"Página 1 de {uploaded_file.name}", use_column_width=True)
                        
                        with col2:
                            # Botão para alternar modo de edição
                            if st.session_state[f"edit_mode_{i}"]:
                                st.markdown('<div class="edit-mode-container">', unsafe_allow_html=True)
                                st.markdown('<div class="slider-label">Modo de edição ativo</div>', unsafe_allow_html=True)
                                
                                # Sliders para coordenadas de referência
                                st.markdown('<div class="slider-label slider-blue">Área de Referência/Instalação (Azul)</div>', unsafe_allow_html=True)
                                ref_x1 = st.slider("X1 (Referência)", 0, img.width, st.session_state[f"temp_ref_coords_{i}"][0], key=f"temp_ref_x1_{i}")
                                ref_y1 = st.slider("Y1 (Referência)", 0, img.height, st.session_state[f"temp_ref_coords_{i}"][1], key=f"temp_ref_y1_{i}")
                                ref_x2 = st.slider("X2 (Referência)", 0, img.width, st.session_state[f"temp_ref_coords_{i}"][2], key=f"temp_ref_x2_{i}")
                                ref_y2 = st.slider("Y2 (Referência)", 0, img.height, st.session_state[f"temp_ref_coords_{i}"][3], key=f"temp_ref_y2_{i}")
                                
                                # Atualizar coordenadas temporárias
                                st.session_state[f"temp_ref_coords_{i}"] = [ref_x1, ref_y1, ref_x2, ref_y2]
                                
                                # Sliders para coordenadas da tabela
                                st.markdown('<div class="slider-label slider-green">Área da Tabela (Verde)</div>', unsafe_allow_html=True)
                                tab_x1 = st.slider("X1 (Tabela)", 0, img.width, st.session_state[f"temp_tab_coords_{i}"][0], key=f"temp_tab_x1_{i}")
                                tab_y1 = st.slider("Y1 (Tabela)", 0, img.height, st.session_state[f"temp_tab_coords_{i}"][1], key=f"temp_tab_y1_{i}")
                                tab_x2 = st.slider("X2 (Tabela)", 0, img.width, st.session_state[f"temp_tab_coords_{i}"][2], key=f"temp_tab_x2_{i}")
                                tab_y2 = st.slider("Y2 (Tabela)", 0, img.height, st.session_state[f"temp_tab_coords_{i}"][3], key=f"temp_tab_y2_{i}")
                                
                                # Atualizar coordenadas temporárias
                                st.session_state[f"temp_tab_coords_{i}"] = [tab_x1, tab_y1, tab_x2, tab_y2]
                                
                                # Botão para aplicar seleção
                                if st.button("Aplicar seleção", key=f"aplicar_{i}"):
                                    # Copiar valores temporários para os valores confirmados
                                    st.session_state[f"ref_coords_{i}"] = st.session_state[f"temp_ref_coords_{i}"].copy()
                                    st.session_state[f"tab_coords_{i}"] = st.session_state[f"temp_tab_coords_{i}"].copy()
                                    st.session_state[f"edit_mode_{i}"] = False
                                    st.experimental_rerun()
                                
                                st.markdown('</div>', unsafe_allow_html=True)
                            else:
                                # Botão para entrar no modo de edição
                                if st.button("Editar seleção", key=f"editar_{i}"):
                                    # Copiar valores confirmados para os valores temporários
                                    st.session_state[f"temp_ref_coords_{i}"] = st.session_state[f"ref_coords_{i}"].copy()
                                    st.session_state[f"temp_tab_coords_{i}"] = st.session_state[f"tab_coords_{i}"].copy()
                                    st.session_state[f"edit_mode_{i}"] = True
                                    st.experimental_rerun()
                                
                                # Exibir coordenadas atuais
                                st.markdown('<div class="slider-label slider-blue">Área de Referência/Instalação (Azul)</div>', unsafe_allow_html=True)
                                st.write(f"X1: {st.session_state[f'ref_coords_{i}'][0]}, Y1: {st.session_state[f'ref_coords_{i}'][1]}")
                                st.write(f"X2: {st.session_state[f'ref_coords_{i}'][2]}, Y2: {st.session_state[f'ref_coords_{i}'][3]}")
                                
                                st.markdown('<div class="slider-label slider-green">Área da Tabela (Verde)</div>', unsafe_allow_html=True)
                                st.write(f"X1: {st.session_state[f'tab_coords_{i}'][0]}, Y1: {st.session_state[f'tab_coords_{i}'][1]}")
                                st.write(f"X2: {st.session_state[f'tab_coords_{i}'][2]}, Y2: {st.session_state[f'tab_coords_{i}'][3]}")
                                
                                # Botão para processar esta fatura
                                if st.button("Processar esta fatura", key=f"processar_{i}"):
                                    try:
                                        # Converter coordenadas da imagem para coordenadas do PDF
                                        with pdfplumber.open(io.BytesIO(pdf_desbloqueado)) as pdf:
                                            if len(pdf.pages) > 0:
                                                page = pdf.pages[0]
                                                pdf_width = page.width
                                                pdf_height = page.height
                                                
                                                # Fator de escala
                                                scale_x = pdf_width / img.width
                                                scale_y = pdf_height / img.height
                                                
                                                # Coordenadas ajustadas
                                                ref_coords = [
                                                    int(st.session_state[f"ref_coords_{i}"][0] * scale_x),
                                                    int(st.session_state[f"ref_coords_{i}"][1] * scale_y),
                                                    int(st.session_state[f"ref_coords_{i}"][2] * scale_x),
                                                    int(st.session_state[f"ref_coords_{i}"][3] * scale_y)
                                                ]
                                                
                                                tab_coords = [
                                                    int(st.session_state[f"tab_coords_{i}"][0] * scale_x),
                                                    int(st.session_state[f"tab_coords_{i}"][1] * scale_y),
                                                    int(st.session_state[f"tab_coords_{i}"][2] * scale_x),
                                                    int(st.session_state[f"tab_coords_{i}"][3] * scale_y)
                                                ]
                                                
                                                # Salvar padrão de seleção
                                                salvar_padrao_selecao(distribuidora, ref_coords, tab_coords)
                                                
                                                # Extrair texto das áreas selecionadas
                                                texto_ref = extrair_texto_area(pdf_desbloqueado, 0, *ref_coords)
                                                texto_tab = extrair_tabela_area(pdf_desbloqueado, 0, *tab_coords)
                                                
                                                # Processar com OpenAI
                                                dados, tokens, custo = processar_com_openai(api_key, texto_tab, modelo)
                                                
                                                # Atualizar contadores
                                                st.session_state.total_tokens += tokens
                                                st.session_state.total_custo += custo
                                                
                                                # Classificar itens
                                                if dados:
                                                    classificacao = classificar_itens(dados, dicionario)
                                                    
                                                    # Extrair informações de referência e instalação
                                                    padrao_instalacao = r'Instalação:\s*(\d+)'
                                                    padrao_referencia = r'Referente a\s*([A-Za-z]+/\d{4})'
                                                    
                                                    instalacao = re.search(padrao_instalacao, texto_ref)
                                                    instalacao = instalacao.group(1) if instalacao else "N/A"
                                                    
                                                    referencia = re.search(padrao_referencia, texto_ref)
                                                    referencia = referencia.group(1) if referencia else "N/A"
                                                    
                                                    # Adicionar aos resultados
                                                    st.session_state.resultados.append({
                                                        "Arquivo": uploaded_file.name,
                                                        "Distribuidora": distribuidora,
                                                        "Instalação": instalacao,
                                                        "Referência": referencia,
                                                        "Dados": dados,
                                                        "Classificação": classificacao,
                                                        "Tokens": tokens,
                                                        "Custo": custo
                                                    })
                                                    
                                                    st.success(f"Fatura processada com sucesso! Custo: ${custo:.6f}")
                                                else:
                                                    st.error("Não foi possível extrair dados estruturados. Tente ajustar a seleção da tabela.")
                                    except Exception as e:
                                        st.error(f"Erro ao processar fatura: {str(e)}")
                    else:
                        st.error(f"Não foi possível converter o PDF {uploaded_file.name} para imagem")
                
                # Atualizar barra de progresso
                progress_bar.progress((i + 1) / len(uploaded_files))
            
            status_text.text("Processamento concluído!")
            progress_bar.progress(1.0)

# Exibir resultados
if st.session_state.resultados:
    st.header("Resultados")
    
    # Exibir custo total
    st.markdown(f"""
    <div class="cost-display">
        <p><strong>Total de tokens processados:</strong> {st.session_state.total_tokens}</p>
        <p><strong>Custo total em USD:</strong> ${st.session_state.total_custo:.6f}</p>
        <p><strong>Custo total em BRL (taxa 1:6):</strong> R${st.session_state.total_custo * 6:.2f}</p>
    </div>
    """, unsafe_allow_html=True)
    
    # Criar DataFrame consolidado
    dados_consolidados = []
    
    for resultado in st.session_state.resultados:
        # Dados básicos
        linha = {
            "Arquivo": resultado["Arquivo"],
            "Distribuidora": resultado["Distribuidora"],
            "Instalação": resultado["Instalação"],
            "Referência": resultado["Referência"],
            "Encargo": resultado["Classificação"]["Encargo"],
            "Desc. Encargo": resultado["Classificação"]["Desc. Encargo"],
            "Demanda": resultado["Classificação"]["Demanda"],
            "Desc. Demanda": resultado["Classificação"]["Desc. Demanda"],
            "Reativa": resultado["Classificação"]["Reativa"],
            "Contb. Publica": resultado["Classificação"]["Contb. Publica"],
            "Outros": resultado["Classificação"]["Outros"],
            "Tokens": resultado["Tokens"],
            "Custo (USD)": resultado["Custo"]
        }
        
        dados_consolidados.append(linha)
    
    # Criar DataFrame
    df_consolidado = pd.DataFrame(dados_consolidados)
    
    # Formatar colunas monetárias
    colunas_monetarias = ["Encargo", "Desc. Encargo", "Demanda", "Desc. Demanda", "Reativa", "Contb. Publica", "Outros"]
    for col in colunas_monetarias:
        df_consolidado[col] = df_consolidado[col].apply(lambda x: f"R$ {x:.2f}".replace(".", ","))
    
    # Formatar coluna de custo
    df_consolidado["Custo (USD)"] = df_consolidado["Custo (USD)"].apply(lambda x: f"$ {x:.6f}")
    
    # Exibir tabela
    st.dataframe(df_consolidado)
    
    # Botão para exportar para Excel
    excel_data = to_excel(df_consolidado)
    st.download_button(
        label="📥 Exportar para Excel",
        data=excel_data,
        file_name=f"faturas_processadas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    
    # Exibir detalhes de cada fatura
    st.header("Detalhes das Faturas")
    
    for i, resultado in enumerate(st.session_state.resultados):
        with st.expander(f"{resultado['Arquivo']} - {resultado['Distribuidora']} - {resultado['Instalação']}"):
            st.subheader("Informações Básicas")
            st.write(f"**Instalação:** {resultado['Instalação']}")
            st.write(f"**Referência:** {resultado['Referência']}")
            st.write(f"**Distribuidora:** {resultado['Distribuidora']}")
            
            st.subheader("Itens Extraídos")
            if resultado["Dados"] and "itens" in resultado["Dados"]:
                # Criar DataFrame com os itens
                itens_df = pd.DataFrame(resultado["Dados"]["itens"])
                st.dataframe(itens_df)
            else:
                st.warning("Nenhum item extraído")
            
            st.subheader("Classificação")
            # Criar DataFrame com a classificação
            classificacao_df = pd.DataFrame([resultado["Classificação"]])
            st.dataframe(classificacao_df)
            
            st.subheader("Métricas")
            st.write(f"**Tokens processados:** {resultado['Tokens']}")
            st.write(f"**Custo em USD:** ${resultado['Custo']:.6f}")
            st.write(f"**Custo em BRL:** R${resultado['Custo'] * 6:.2f}")
