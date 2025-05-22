import streamlit as st
import pandas as pd
import json
import os
import io
import re
import unicodedata
from datetime import datetime
import warnings
import logging
import sys
import contextlib
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import requests
import pymupdf as fitz
import pikepdf
from openai import OpenAI

# Configuração da página Streamlit
st.set_page_config(
    page_title="PDFator - Extrator de Faturas de Energia",
    page_icon="⚡",
    layout="wide"
)

# Estilo CSS personalizado
st.markdown("""
<style>
    .main-header {
        font-size: 2.5rem;
        color: #333;
        margin-bottom: 1rem;
    }
    .sub-header {
        font-size: 1.2rem;
        color: #666;
        margin-bottom: 2rem;
    }
    .info-box {
        background-color: #f8f9fa;
        padding: 1.5rem;
        border-radius: 10px;
        margin-bottom: 1.5rem;
    }
    .stats-container {
        display: flex;
        justify-content: space-between;
        margin-bottom: 1rem;
    }
    .stat-box {
        background-color: #f1f3f5;
        padding: 1rem;
        border-radius: 8px;
        text-align: center;
        flex: 1;
        margin: 0 0.5rem;
    }
    .footer {
        margin-top: 3rem;
        text-align: center;
        color: #888;
        font-size: 0.8rem;
    }
</style>
""", unsafe_allow_html=True)

# Função para ocultar saída
@contextlib.contextmanager
def ocultar_output():
    with open(os.devnull, "w") as fnull:
        old_stdout = sys.stdout
        old_stderr = sys.stderr
        sys.stdout = fnull
        sys.stderr = fnull
        try:
            yield
        finally:
            sys.stdout = old_stdout
            sys.stderr = old_stderr

# ===================================================================
# INSIRA AQUI O MAPEAMENTO DE COBRANÇA
# Cole o conteúdo da variável mapeamento_tipo_cobranca abaixo
# Mantenha o nome da variável como "mapeamento_tipo_cobranca"
# ===================================================================


mapeamento_tipo_cobranca = {
#===================================================================ENCARGO================================================================================================================
        "Energia elet. adquirida 3": "Encargo",
        "energia elet adquirida 3": "Encargo",
        "energiaelet.adquirida3": "Encargo",
        "energia elet. adquirida3": "Encargo",
        "energia elet. adquirida 3* c/imposto": "Encargo",
        "energia elet. adquirida 3* c/lmposto": "Encargo",
        "energia elet. adquirida 3* c imp": "Encargo",
        "energia elet adquirida 3 c imposto": "Encargo",
        "Liminar Red Icms Fatura": "Encargo",
        "liminar red icms fatura": "Encargo",
        "liminarredicmsfatura": "Encargo",
        "liminar red. icms fatura": "Encargo",
        "Subtotal Faturamento": "Encargo",
        "Componente Encargo kWh HFP": "Encargo",
        "componente encargo kwh hfp": "Encargo",
        "Componente Encargo kWh HP": "Encargo",
        "componente encargo kwh hp": "Encargo",
        "Energia Terc Comercializad HFP": "Encargo",
        "energia terc comercializad hfp": "Encargo",
        "Energia Terc Comercializad HP": "Encargo",
        "energia terc comercializad hp": "Encargo",
        "Energia Isenta Comercial. HFP": "Encargo",
        "energia isenta comercial. hfp": "Encargo",
        "PIS/COFINS da subvenção/descon": "Encargo",
        "pis/cofins da subvenção/descon": "Encargo",
        "subtotal faturamento": "Encargo",
        "sub total faturamento": "Encargo",
        "subtot. faturamento": "Encargo",
        "Componente Encargo kWh HFP": "Encargo",
        "componente encargo kwh hfp": "Encargo",
        "componente encargokwh hfp": "Encargo",
        "componente encargo kw hfp": "Encargo",
        "Componente Encargo kWh HP": "Encargo",
        "componente encargo kwh hp": "Encargo",
        "componente encargokwh hp": "Encargo",
        "componente encargo kw hp": "Encargo",
        "Componente Encargo HFP": "Encargo",
        "componente encargo hfp": "Encargo",
        "Componente Encargo HP": "Encargo",
        "componente encargo hp": "Encargo",
        "ENERGIA ACL": "Encargo",
        "energia acl": "Encargo",
        "energiaacl": "Encargo",
        "energia acl ": "Encargo",
        "CONSUMOATIVO PONTA TUSD": "Encargo",
        "consumoativo ponta tusd": "Encargo",
        "consumo ativo ponta tusd": "Encargo",
        "consumo ativoponta tusd": "Encargo",
        "CONSUMO ATIVO F.PONTA TUSD": "Encargo",
        "consumo ativo f.ponta tusd": "Encargo",
        "consumo ativo f. ponta tusd": "Encargo",
        "consumo ativo f ponta tusd": "Encargo",
        "BENEFICIO TARIFARIO BRUTO": "Encargo",
        "beneficio tarifario bruto": "Encargo",
        "benefício tarifario bruto": "Encargo",
        "beneficio tar. bruto": "Encargo",
        "benefício tar. bruto": "Encargo",
        "PIS/COFINS da subvencao/descon": "Encargo",
        "PIS/COFINS da subvencäo/descon": "Encargo",
        "pis cofins da subvençao/descon": "Encargo",
        "piscofins da subvencao descon": "Encargo",
        "Correcao IPCA/IGPM s/ conta 01/25 pg 19/02/25": "Encargo",
        "correção ipca igpm s/ conta": "Encargo",
        "correcao ipca igpm s/conta": "Encargo",
        "correcao ipca/igpm s/conta": "Encargo",
        "correcao ipca igpm s conta": "Encargo",
        "Componente Encargo kWh HFP": "Encargo",
        "componente encargo kwh hfp": "Encargo",
        "componente encargokwh hfp": "Encargo",
        "Componente Encargo kWh HP": "Encargo",
        "componente encargo kwh hp": "Encargo",
        "componente encargokwh hp": "Encargo",
        "PIS/COFINS da subvencao/descon": "Encargo",
        "PIS/COFINS da subvencäo/descon": "Encargo",
        "PIS/COFINS da subvenção/descon": "Encargo",
        "pis cofins da subvençao descon": "Encargo",
        "Encargo Cta Covid REN 885/2020": "Encargo",
        "encargo cta covid ren885/2020": "Encargo",
        "encargo covid ren 885/2020": "Encargo",
        "encargo covid ren885/2020": "Encargo",
        "Energia Ativa kWh HFP/nico": "Encargo",
        "energia ativa kwh hfp/unico": "Encargo",
        "energia ativa kwh hfp unico": "Encargo",
        "energia ativa kwh hfp": "Encargo",
        "Energia Ativa kWh HP": "Encargo",
        "energia ativa kwh hp": "Encargo",
        "DEBITO VAR IPCA": "Encargo",
        "débito var ipca": "Encargo",
        "debito var ipca": "Encargo",
        "deb1to var ipca": "Encargo",
        "ENERGIA ACL": "Encargo",
        "energia acl": "Encargo",
        "energiaacl": "Encargo",
        "TUSD ponta": "Encargo",
        "TUSD fora ponta": "Encargo",
        "tusd ponta": "Encargo",
        "tusd fora ponta": "Encargo",
        "Consumo Energia Elétrica HP": "Encargo",
        "consumo energia eletrica hp": "Encargo",
        "consumo energia elétrica hp": "Encargo",
        "DEMANDALEIESTADUAL16.886/18": "Encargo",




#===================================================================Desc. Encargo===================================================================
        "DEDUCAO ENERGIA ACL": "Desc. Encargo",
        "DEDUCÃO ENERGIA ACL": "Desc. Encargo",
        "DED. ENERGIA ACL": "Desc. Encargo",
        "D. ENERGIA ACL": "Desc. Encargo",
        "Valor de Autoprodução HFP": "Desc. Encargo",
        "BENEFICIO TARIFARIO LIQUIDO": "Desc. Encargo",
        "BENEFÍCIO TARIFÁRIO LIQUIDO": "Desc. Encargo",
        "BENEFiCIO TARIFARIO LIQUIDO": "Desc. Encargo",
        "BENEFÍCIO TARIFÁRIO LÍQUIDO": "Desc. Encargo",
        "BENEFICIO TARIFÁRIO LÍQUIDO": "Desc. Encargo",
        "liminar icm5 tusd-consumo": "Desc. Encargo",
        "liminar icms tusd-consu mo": "Desc. Encargo",
        "liminar icms tusd-con5umo": "Desc. Encargo",
        "liminaricms tusd consumo": "Desc. Encargo",
        "liminar icms t usd-consumo": "Desc. Encargo",
        "liminar icms tu sd consumo": "Desc. Encargo",
        "liminar icms tusd-con sumo": "Desc. Encargo",
        "liminar icms t u s d-consumo": "Desc. Encargo",
        "liminar icms tusd-demanda": "Desc. Encargo",
        "liminar icms tusd demanda": "Desc. Encargo",
        "liminar icms tusddemanda": "Desc. Encargo",
        "liminar icms tus ddemanda": "Desc. Encargo",
        "liminar 1cms tusd-demanda": "Desc. Encargo",
        "liminar icms tuso-demanda": "Desc. Encargo",
        "liminar icm5 tusd-demanda": "Desc. Encargo",
        "liminar icms tusd demand a": "Desc. Encargo",
        "liminaricms tusd demanda": "Desc. Encargo",
        "liminar icms tu sd-demanda": "Desc. Encargo",
        "liminar icms t usd-demanda": "Desc. Encargo",
        "liminar icms tus-d-demanda": "Desc. Encargo",
        "liminar icms tusd d emanda": "Desc. Encargo",
        "liminar icms t u s d-demanda": "Desc. Encargo",
        "Liminar ICMS TUSD-Demanda": "Desc. Encargo",
        "Liminar ICMS TUSD Demanda": "Desc. Encargo",
        "Liminar ICMS TUSDDemanda": "Desc. Encargo",
        "Liminar ICMS TUS D Demanda": "Desc. Encargo",
        "Liminar ICMS TUS-D-Demanda": "Desc. Encargo",
        "Liminar ICMS TUSD Deman da": "Desc. Encargo",
        "Liminar 1CMS TUSD-Demanda": "Desc. Encargo",
        "Llminar ICMS TUSD-Demanda": "Desc. Encargo",
        "Liminar ICMS TUSO-Demanda": "Desc. Encargo",
        "Liminar ICMS TUS-Demanda": "Desc. Encargo",
        "Liminar ICMS TUSDDemand a": "Desc. Encargo",
        "Liminar ICMS TUSD-Consumo": "Desc. Encargo",
        "Liminar ICMS TUSD Consumo": "Desc. Encargo",
        "DEDUCAO ENERGIAACL": "Desc. Encargo",
        "DEDUÇÃO ENERGIAACL": "Desc. Encargo",
        "DED. ENERGIAACL": "Desc. Encargo",
        "deducao energiaacl": "Desc. Encargo",
        "deducaoenergiaacl":  "Desc. Encargo",
        "deducao energia acl": "Desc. Encargo",
        "DEDUÇÃO ENERGIA ACL": "Desc. Encargo",
        "Liminar ICMS TUSDConsumo": "Desc. Encargo",
        "Liminar ICMS TUS D Consumo": "Desc. Encargo",
        "Liminar ICMS TUSD-Consu mo": "Desc. Encargo",
        "Liminar ICMS TUSO-Consumo": "Desc. Encargo",
        "Liminar 1CMS TUSD-Consumo": "Desc. Encargo",
        "Llminar ICMS TUSD-Consumo": "Desc. Encargo",
        "Liminar ICMS TUSD-Consuno": "Desc. Encargo",
        "Liminar ICMS TUS DConsumo": "Desc. Encargo",
        "DEDUCAO ENERGIA ACL": "Desc. Encargo",
        "DEDUCAO ENERGIA ACL ": "Desc. Encargo",
        "Valor de Autoproducao FP": "Desc. Encargo",
        "DEDUCAO ENERGIAACL": "Desc. Encargo",
        "DEDUTCAO ENERGIA ACL": "Desc. Encargo",
        "DEDUCAO ENERGIA ACl": "Desc. Encargo",
        "DEDUCAO ENERGIA ACL.": "Desc. Encargo",
        "DEDUTCAO ENERGIA ACL ": "Desc. Encargo",
        "DEDULCAO ENERGIA ACL": "Desc. Encargo",
        "DED0CAO ENERGIA ACL": "Desc. Encargo",
        "DEDUCAO ENERGIA  ACL": "Desc. Encargo",
        "DEDUC0AO ENERGIA ACL": "Desc. Encargo",
        "DEDUCAO ENERGIA_ACL": "Desc. Encargo",
        "DED0CAO ENERGIA_ACL": "Desc. Encargo",
        "DEDUCAO ENERGIA ACL-": "Desc. Encargo",
        "DEDUCAO ENERGIA ACL :": "Desc. Encargo",
        "DEDUCAO ENERGIA ACL .": "Desc. Encargo",
        "DEDUCAO ENERGIA ACL,": "Desc. Encargo",
        "Valor de Autoprodução HP": "Desc. Encargo",
        "BENEFiCIO TARIFARIO LIQUIDO": "Desc. Encargo",
        "BENEFÍCIO TARIFÁRIO LIQUIDO": "Desc. Encargo",
        "BENEFÍCIO TARIFARIO LIQUIDO": "Desc. Encargo",
        "BENEFICIO TARIFARIO LIQUIDO": "Desc. Encargo",
        "DEDUCAO ENERGIA ACL": "Desc. Encargo",
        "DEDUCAO ENERGIA ACL ": "Desc. Encargo",
        "DEDUCAO ENERGIAACL": "Desc. Encargo",
        "DEDUCÃO ENERGIA ACL": "Desc. Encargo",
        "DEDUCAO ENERGIAAC": "Desc. Encargo",
        "DEDUCAO ENERGIAA CL": "Desc. Encargo",
        "DEDUCAO ENERGIA-ACL": "Desc. Encargo",
        "DEDUCAO-ENERGIA ACL": "Desc. Encargo",
        "DEDUC0AO ENERGIA ACL": "Desc. Encargo",
        "DEDUCAOENERGIA ACL": "Desc. Encargo",
        "DEDUTCAO ENERGIA ACL": "Desc. Encargo",
        "DEDUCAO ENERGIA ACL    ": "Desc. Encargo",
        "DEDUCAO ENERGIAACL.": "Desc. Encargo",
        "DEDUCAO ENERGIAACL:": "Desc. Encargo",
        "DED0CAO ENERGIA ACL": "Desc. Encargo",
        "DEDUCAO ENERGIA A C L": "Desc. Encargo",
        "DEDUCAO ENERGIAACL-": "Desc. Encargo",
        "DEDUCAO ENERGIAA-CL": "Desc. Encargo",
        "DEDUCAO ENERGIAACL ,": "Desc. Encargo",
        "Desconto Comp. Encargo HP": "Desc. Encargo",
        "Ajuste de Desconto C. Enc HP": "Desc. Encargo",
        "Desconto Comp. Encargo HFP": "Desc. Encargo",
        "Ajuste de Desconto C. Enc HFP": "Desc. Encargo",
        "desconto compencargo hp": "Desc. Encargo",
        "ajuste de desconto cenc hp": "Desc. Encargo",
        "Valor Autoprodução HFP": "Desc. Encargo",
        "Valor de Autoproducao HFP": "Desc. Encargo",
        "Valor Autoprodução HP": "Desc. Encargo",
        "Valor de Autoproducao HP": "Desc. Encargo",
        "Desconto Comp Encargo HP": "Desc. Encargo",
        "Desconto Compencargo HP": "Desc. Encargo",
        "Ajuste de Desconto C Enc HP": "Desc. Encargo",
        "Ajuste de Desconto CEnc HP": "Desc. Encargo",
        "Ajuste de Desconto C Enc HFP": "Desc. Encargo",
        "Ajuste de Desconto CEnc HFP": "Desc. Encargo",
        "Desconto Comp Encargo HFP": "Desc. Encargo",
        "Desconto Compencargo HFP": "Desc. Encargo",
        "Valor de Autoproducao FP": "Desc. Encargo",
        "Valor de Autoprodução FP": "Desc. Encargo",
        "Valor de Autoproduç:ao FP": "Desc. Encargo",
        "Valor de Autoproducäo FP": "Desc. Encargo",
        "Valor Autoproducao FP": "Desc. Encargo",
        "Valor de Autoproducao F P": "Desc. Encargo",
        "Autoproducao FP": "Desc. Encargo",
        "Valor Autoproducao": "Desc. Encargo",
        "Restituicao de Pagamento": "Desc. Encargo",
        "Restituição de Pagamento": "Desc. Encargo",
        "Restituic:ao de Pagamento": "Desc. Encargo",
        "Restitüição de Pagamento": "Desc. Encargo",
        "Restituicaoo de Pagamento": "Desc. Encargo",
        "Restituicão Pagamento": "Desc. Encargo",
        "liminar icms tusd-consumo": "Desc. Encargo",
        "liminar icms tusd consumo": "Desc. Encargo",
        "liminar icms tusdconsumo": "Desc. Encargo",
        "liminar icms tus dconsumo": "Desc. Encargo",
        "liminar 1cms tusd-consumo": "Desc. Encargo",
        "liminar icm5 tusd-consumo": "Desc. Encargo",
        "liminar icms tusd-consu mo": "Desc. Encargo",
        "liminar icms tusd-con5umo": "Desc. Encargo",
        "liminaricms tusd consumo": "Desc. Encargo",
        "liminar icms t usd-consumo": "Desc. Encargo",
        "liminar icms tu sd consumo": "Desc. Encargo",
        "liminar icms tusd-con sumo": "Desc. Encargo",
        "liminar icms t u s d-consumo": "Desc. Encargo",
        "liminar icms tusd-demanda": "Desc. Encargo",
        "liminar icms tusd demanda": "Desc. Encargo",
        "liminar icms tusddemanda": "Desc. Encargo",
        "liminar icms tus ddemanda": "Desc. Encargo",
        "liminar 1cms tusd-demanda": "Desc. Encargo",
        "liminar icms tuso-demanda": "Desc. Encargo",
        "liminar icm5 tusd-demanda": "Desc. Encargo",
        "liminar icms tusd demand a": "Desc. Encargo",
        "liminaricms tusd demanda": "Desc. Encargo",
        "liminar icms tu sd-demanda": "Desc. Encargo",
        "liminar icms t usd-demanda": "Desc. Encargo",
        "liminar icms tus-d-demanda": "Desc. Encargo",
        "liminar icms tusd d emanda": "Desc. Encargo",
        "liminar icms t u s d-demanda": "Desc. Encargo",
        "Liminar ICMS TUSD-Demanda": "Desc. Encargo",
        "Liminar ICMS TUSD Demanda": "Desc. Encargo",
        "Liminar ICMS TUSDDemanda": "Desc. Encargo",
        "Liminar ICMS TUS D Demanda": "Desc. Encargo",
        "Liminar ICMS TUS-D-Demanda": "Desc. Encargo",
        "Liminar ICMS TUSD Deman da": "Desc. Encargo",
        "Liminar 1CMS TUSD-Demanda": "Desc. Encargo",
        "Llminar ICMS TUSD-Demanda": "Desc. Encargo",
        "Liminar ICMS TUSO-Demanda": "Desc. Encargo",
        "Liminar ICMS TUS-Demanda": "Desc. Encargo",
        "Liminar ICMS TUSDDemand a": "Desc. Encargo",
        "Liminar ICMS TUSD-Consumo": "Desc. Encargo",
        "Liminar ICMS TUSD Consumo": "Desc. Encargo",
        "Liminar ICMS TUSDConsumo": "Desc. Encargo",
        "Liminar ICMS TUS D Consumo": "Desc. Encargo",
        "Liminar ICMS TUSD-Consu mo": "Desc. Encargo",
        "Liminar ICMS TUSO-Consumo": "Desc. Encargo",
        "Liminar 1CMS TUSD-Consumo": "Desc. Encargo",
        "Llminar ICMS TUSD-Consumo": "Desc. Encargo",
        "Liminar ICMS TUSD-Consuno": "Desc. Encargo",
        "Liminar ICMS TUS DConsumo": "Desc. Encargo",
        "Valor de Autoproducao FP": "Desc. Encargo",
        "Valor de Autoprodução FP": "Desc. Encargo",
        "Valor de Autoproduç:ao FP": "Desc. Encargo",
        "Valor de Autoproducäo FP": "Desc. Encargo",
        "Valor Autoproducao FP": "Desc. Encargo",
        "Valor de Autoproducao F P": "Desc. Encargo",
        "Autoproducao FP": "Desc. Encargo",
        "Valor Autoproducao": "Desc. Encargo",
        "Restituicao de Pagamento": "Desc. Encargo",
        "Restituição de Pagamento": "Desc. Encargo",
        "Restituic:ao de Pagamento": "Desc. Encargo",
        "Restitüição de Pagamento": "Desc. Encargo",
        "Restituicaoo de Pagamento": "Desc. Encargo",
        "Restituicão Pagamento": "Desc. Encargo",
        "Valor de Autoprodução FP": "Desc. Encargo",
        "Valor de Autoproducao FP": "Desc. Encargo",
        "Valor de Autoproducão FP": "Desc. Encargo",
        "Valor de Autoproducäo FP": "Desc. Encargo",
        "Valor de Autoprodução HFP": "Desc. Encargo",
        "Valor de Autoproducao HFP": "Desc. Encargo",
        "Valor de Autoproducão HFP": "Desc. Encargo",
        "Valor de Autoproducäo HFP": "Desc. Encargo",
        "Valor de Autoprodução HP": "Desc. Encargo",
        "Valor de Autoproducao HP": "Desc. Encargo",
        "Valor Autoproducao FP": "Desc. Encargo",
        "Valor Autoproducao HFP": "Desc. Encargo",
        "Valor Autoproducao HP": "Desc. Encargo",
        "DEDUCAO ENERGIA ACL": "Desc. Encargo",
        "DEDUÇÃO ENERGIA ACL": "Desc. Encargo",
        "DEDUCAO ENERGIAACL": "Desc. Encargo",
        "DEDUCAO ENERGIAA CL": "Desc. Encargo",
        "DEDUCAOENERGIA ACL": "Desc. Encargo",
        "DEDUCÃO ENERGIA ACL": "Desc. Encargo",
        "BENEFICIO TARIFARIO LIQUIDO": "Desc. Encargo",
        "BENEFÍCIO TARIFÁRIO LIQUIDO": "Desc. Encargo",
        "BENEFiCIO TARIFARIO LIQUIDO": "Desc. Encargo",
        "BENEFÍCIO TARIFÁRIO LÍQUIDO": "Desc. Encargo",
        "Desconto Comp.Encargo HP": "Desc. Encargo",
        "Desconto Comp Encargo HP": "Desc. Encargo",
        "Desconto Compencargo HP": "Desc. Encargo",
        "Ajuste de Desconto C.Enc HP": "Desc. Encargo",
        "Ajuste de Desconto C Enc HP": "Desc. Encargo",
        "Ajuste de Desconto CEnc HP": "Desc. Encargo",
        "Desconto Comp.Encargo HFP": "Desc. Encargo",
        "Ajuste de Desconto C.Enc HFP": "Desc. Encargo",
        "Ajuste de Desconto C Enc HFP": "Desc. Encargo",
        "Ajuste de Desconto CEnc HFP": "Desc. Encargo",
        "Restituicao de Pagamento": "Desc. Encargo",
        "Restituição de Pagamento": "Desc. Encargo",
        "TAR.REDUZ.RES.ANEEL166/05P": "Desc. Encargo",
        "TAR.REDUZ.RES.ANEEL166/05FP": "Desc. Encargo",
        "Diferenca Energia Retroativa": "Desc. Encargo",
        "Desconto Comp.Encargo HP": "Desc. Encargo",
        "Ajuste de Desconto C.Enc HP": "Desc. Encargo",
        "Ajuste de Desconto C.Enc HFP": "Desc. Encargo",
        "Valor de Autoprodução FP": "Desc. Encargo",
        "Valor de Autoproducao FP": "Desc. Encargo",
        "Valor de Autoprodução HFP": "Desc. Encargo",
        "Valor de Autoproducao HFP": "Desc. Encargo",
        "Valor de Autoprodução HP": "Desc. Encargo",
        "Valor de Autoproducao HP": "Desc. Encargo",
        "DEDUCAO ENERGIA ACL": "Desc. Encargo",
        "DEDUCAO ENERGIAACL": "Desc. Encargo",
        "BENEFICIO TARIFARIO LIQUIDO": "Desc. Encargo",
        "TAR.REDUZ.RES.ANEEL166/05P": "Desc. Encargo",
        "TAR.REDUZ.RES.ANEEL166/05FP": "Desc. Encargo",
        "Restituicao de Pagamento": "Desc. Encargo",
        "Diferenca Energia Retroativa": "Desc. Encargo",
        "Energia Terc Comercializad HP": "Desc. Encargo",
        "Energia Terc Comercializad HFP": "Desc. Encargo",

#===================================================================Demanda===================================================================
        "demanda ponta c/desconto": "Demanda",
        "demanda ponta cidesconto": "Demanda",
        "demanda ponta cidescon": "Demanda",
        "demanda ponta c/desc0nt0": "Demanda",
        "DEMANDA FORA PONTA C/DESC": "Demanda",
        "DEMANDA FORA PONTA C/DESC0": "Demanda",
        "DEMANDA FORA PONTA C/DESCONTO": "Demanda",
        "DEMANDA FORA PONTA C/DESC0NT0": "Demanda",
        "DEMANDA FORA PONTA CIDESCONTO": "Demanda",
        "DEMANDA FORA PONTA CIDESC0NTO": "Demanda",
        "Componente Fio kW HFP": "Demanda",
        "componente fio kw hfp": "Demanda",
        "Componente Fio kW HP": "Demanda",
        "componente fio kw hp": "Demanda",
        "demanda p0nta c/desconto": "Demanda",
        "demanda ponta c / desconto": "Demanda",
        "ULTRAPASSAGEM PONTA/TUSD": "Demanda",
        "ULTRAPASSAGEM PONTA TUSD": "Demanda",
        "ULTRAPASSAGEM PONTA TUSD-": "Demanda",
        "ULTRAPASSAGEM PONTA/TUSD ": "Demanda",
        "ULTRAPASSAGEM PONTA/T USD": "Demanda",
        "ULTRAPASSAGEM PONTA/TUSD-": "Demanda",
        "ULTRAPASSAGEM PONTA/TUSD  ": "Demanda",
        "ULTRAPASSAGEM PONTA-TUSD": "Demanda",
        "ULTRA PASSAGEM PONTA/TUSD": "Demanda",
        "ULTRA PASSAGEM PONTA TUSD": "Demanda",
        "ULTRAPASSAGEM PONTA TUSD ": "Demanda",
        "ULTRAPASSAGEM FORAPONTA/TUSD": "Demanda",
        "ULTRAPASSAGEM FORA PONTA TUSD": "Demanda",
        "ULTRAPASSAGEM FORAPONTA TUSD": "Demanda",
        "ULTRAPASSAGEM FORA PONTA/TUSD": "Demanda",
        "ULTRAPASSAGEM FORA PONTA TUSD ": "Demanda",
        "UFDR PONTA/TUSD": "Demanda",
        "UFDRPONTA/TUSD": "Demanda",
        "UFDR PONTA TUSD": "Demanda",
        "UFDRPONTATUSD": "Demanda",
        "UFDR P0NTA/TUSD": "Demanda",
        "UFDR PONTA/T USD": "Demanda",
        "UFDR P0NTA TUSD": "Demanda",
        "UFDR PONTA-TUSD": "Demanda",
        "UFDR PONTA/TUS D": "Demanda",
        "UFDRFORA PONTA/TUSD": "Demanda",
        "UFDR FORA PONTA/TUSD": "Demanda",
        "UFDR FORAPONTA/TUSD": "Demanda",
        "UFDR FORA PONTA TUSD": "Demanda",
        "UFDRFORAPONTA TUSD": "Demanda",
        "UFDRFORA PONTA TUSD": "Demanda",
        "UFDR FORA P0NTA/TUSD": "Demanda",
        "UFDRFORA PONTA-TUSD": "Demanda",
        "UFDR FORA PONTA/TUS D": "Demanda",
        "ULTRAPASSAGEM FORA PONTA T USD": "Demanda",
        "ULTRAPASSAGEM FORA PONTA TUSD-": "Demanda",
        "ULTRAPASSAGEM FORA PONTA/TUSD ": "Demanda",
        "ULTRAPASSAGEMFORAPONTA TUSD": "Demanda",
        "ULTRAPASSAGEMF.PONTA/TUSD": "Demanda",
        "ULTRAPASSAGEM FORA PONTA/TUSD-": "Demanda",
        "ULTRA PASSAGEM FORA PONTA/TUSD": "Demanda",
        "demanda pontac/desconto": "Demanda",
        "demanda ponta c/desconto": "Demanda",
        "demanda ponta cidesconto": "Demanda",
        "demanda ponta cidescon": "Demanda",
        "demanda ponta c/desc0nt0": "Demanda",
        "Componente Fio kW HFP": "Demanda",
        "Componente Fio kW HP": "Demanda",
        "Component Fio kW HFP": "Demanda",
        "Component Fio kW HP": "Demanda",
        "Componente Fio kWHP": "Demanda",
        "Componente Fio KW HP": "Demanda",
        "Componente Fio KW HFP": "Demanda",
        "Componete Fio kW HFP": "Demanda",
        "Comp Fio kW HFP": "Demanda",
        "Comp Fio kW HP": "Demanda",
        "Componente Fio kw hfp": "Demanda",
        "demanda p0nta c/desconto": "Demanda",
        "demanda ponta c / desconto": "Demanda",
        "demanda f0ra p0nta cidescont0": "Demanda",
        "demanda fora ponta cidesconto": "Demanda",
        "demanda fora p0nta cidesconto": "Demanda",
        "demanda fora ponta cidesc0nto": "Demanda",
        "demanda fora ponta cidescon": "Demanda",
        "demanda f ora ponta cidesconto": "Demanda",
        "demanda fora pontacidesconto": "Demanda",
        "DEMANDA PONTA C/DESCONTO": "Demanda",
        "DEMANDA PONTA C DESCONTO": "Demanda",
        "DEMANDA PONTA C/DESC0NT0": "Demanda",
        "DEMANDA P0NTA C/DESCONTO": "Demanda",
        "DEMANDA P0NTA C DESCONTO": "Demanda",
        "DEMANDA FORA PONTA CIDESCONTO": "Demanda",
        "DEMANDA F0RA P0NTA CIDESCONTO": "Demanda",
        "DEMANDA FORA PONTA C/DESC": "Demanda",
        "DEMANDA  FORA  PONTA  C/DESC": "Demanda",
        "DEMANDA FORA PONTA C/DESC0NT0": "Demanda",
        "Consumo Energia Eletrica HFP": "Demanda",
        "Consumo de Energia Eletrica HFP": "Demanda",
        "Consumo Energia Eletrica HFP/Unico": "Demanda",
        "Consumo de Energia Elétrica HFP": "Demanda",
        "Consumo Energia Elétric HFP": "Demanda",
        "Consumo Energia Eléctrica HFP": "Demanda",
        "Consumo Energia Eletrica kWh HFP": "Demanda",
        "Componente Fio HFP s/ ICMS": "Demanda",
        "Componente Fio HFP": "Demanda",
        "Componente Fio HP": "Demanda",
        "Componente Fio HFP s/ICMS": "Demanda",
        "Componente Fio kW HFP": "Demanda",
        "Componente Fio kW HP": "Demanda",
        "Component Fio kW HFP": "Demanda",
        "Component Fio kW HP": "Demanda",
        "Componente Fio kWHP": "Demanda",
        "Componente Fio KW HP": "Demanda",
        "Componente Fio KW HFP": "Demanda",
        "Componete Fio kW HFP": "Demanda",
        "Comp Fio kW HFP": "Demanda",
        "Comp Fio kW HP": "Demanda",
        "Componente Fio kw hfp": "Demanda",
        "Demanda Ativa kW HFP/Unico": "Demanda",
        "Demanda Ativa kW HFP/unico": "Demanda",
        "Demanda Ativa kW HFP/ünico": "Demanda",
        "Demanda Ativa kW HFP/Un1co": "Demanda",
        "Demanda Ativa kW HFP/Unic0": "Demanda",
        "Demanda AtivakW HFP/Unico": "Demanda",
        "Demanda Ativa kWHFP/Unico": "Demanda",
        "Componente Fio kW HFP": "Demanda",
        "Componente Fio kW HP": "Demanda",
        "Component Fio kW HFP": "Demanda",
        "Component Fio kW HP": "Demanda",
        "Componente Fio kWHP": "Demanda",
        "Componente Fio KW HP": "Demanda",
        "Componente Fio KW HFP": "Demanda",
        "Componete Fio kW HFP": "Demanda",
        "Comp Fio kW HFP": "Demanda",
        "Comp Fio kW HP": "Demanda",
        "Componente Fio kw hfp": "Demanda",
        "Componente Fio HFP": "Demanda",
        "Componente Fio HP": "Demanda",
        "Componente Fio kW HP": "Demanda",
        "Componente Fio HFP s/ICMS": "Demanda",
        "Componente Fio HP s/ICMS": "Demanda",
        "Componente Fio HP s/ ICMS": "Demanda",
        "Componente Fio HFP s/ ICMS": "Demanda",
        "DEMANDA PONTA C/DESCONTO": "Demanda",
        "DEMANDA FORA PONTA CIDESCONTO": "Demanda",
        "DEMANDA UNICA C/DESCONTO": "Demanda",
        "Demanda Distribuicao F. Ponta": "Demanda",
        "Demanda Distribuicao Ponta": "Demanda",
        "Dernanda Distribuicao Ponta": "Demanda",
        "DIF.FATUR.TUSD-HOMOLOG.CCEE": "Demanda",
        "ULTRAPASSAGEM FORAPONTA/TUSD": "Demanda",




#===================================================================DescDemanda===================================================================
        "DIF.FATUR.TUSD-HOMOLOG.CCEE": "Desc. Demanda",
        "DIF.FATUR.TUSD-HOMOLOG.C0EE": "Desc. Demanda",
        "DIF.FATUR.TUSD-HOMOLOG.0CEE": "Desc. Demanda",
        "DIF.FATUR.TUSD HOMOLOG CCEE": "Desc. Demanda",
        "DIF.FATUR TUSD-HOMOLOG CCEE": "Desc. Demanda",
        "DIF.FATUR.T0SD-HOMOLOG.CCEE": "Desc. Demanda",
        "D1F.FATUR.TUSD-H0MOLOG.CCEE": "Desc. Demanda",
        "DIF.FATUR.TUSD HOMOLOG.CCEE": "Desc. Demanda",
        "DIF FATUR.TUSD-HOMOLOG.CCEE": "Desc. Demanda",
        "DIF.FATUR.TUSD-HOMOLOG.C0E": "Desc. Demanda",
        "DIF.FATUR.TUSD-HOMOLOG.CEE": "Desc. Demanda",
        "DIF.FATUR.TUSD-HOMOLOG.CCE": "Desc. Demanda",
        "ajuste de desconto cfio hfp": "Desc. Demanda",
        "Ajuste de Desconto C. Fio HFP": "Desc. Demanda",
        "ajuste de desconto c. fio hfp": "Desc. Demanda",
        "Ajuste de Desconto C. Fio HP": "Desc. Demanda",
        "ajuste de desconto c. fio hp": "Desc. Demanda",
        "Desconto Comp. Fio HFP": "Desc. Demanda",
        "desconto comp. fio hfp": "Desc. Demanda",
        "Desconto Comp. Fio HP": "Desc. Demanda",
        "desconto comp. fio hp": "Desc. Demanda",
        "desconto compfio hfp": "Desc. Demanda",
        "Ajuste de Desconto C Fio HFP": "Desc. Demanda",
        "Ajuste de Desconto CFio HFP": "Desc. Demanda",
        "Ajuste de Desconto cfio hfp": "Desc. Demanda",
        "Desconto Comp Fio HFP": "Desc. Demanda",
        "Desconto Compfio HFP": "Desc. Demanda",
        "Ajuste de Desconto C Fio HP": "Desc. Demanda",
        "Ajuste de Desconto CFio HP": "Desc. Demanda",
        "Desconto Comp Fio HP": "Desc. Demanda",
        "Desconto Compfio HP": "Desc. Demanda",
        "Ajuste de Desconto C. Fio HFP": "Desc. Demanda",
        "Ajuste de Desconto C. Fio HP": "Desc. Demanda",
        "Desconto Comp. Fio HFP": "Desc. Demanda",
        "Desconto Comp. Fio HP": "Desc. Demanda",
        "Desconto Comp.Fio HP": "Desc. Demanda",
        "Desconto Comp Fio HP": "Desc. Demanda",
        "Desc. Comp. Fio HP": "Desc. Demanda",
        "Desc Comp.Fio HP": "Desc. Demanda",
        "Desconto Compfio HP": "Desc. Demanda",
        "Desconto Comp FioHP": "Desc. Demanda",
        "Desconto Componente Fio HP": "Desc. Demanda",
        "Desconto Comp.Fio HFP": "Desc. Demanda",
        "Desconto Comp Fio HFP": "Desc. Demanda",
        "Desc. Comp. Fio HFP": "Desc. Demanda",
        "Desc Comp.Fio HFP": "Desc. Demanda",
        "Desconto Compfio HFP": "Desc. Demanda",
        "Desconto Comp FioHFP": "Desc. Demanda",
        "Desconto Componente Fio HFP": "Desc. Demanda",
        "Ajuste de Desconto C.Fio HP": "Desc. Demanda",
        "Ajuste de Desconto C. Fio HP": "Desc. Demanda",
        "Ajuste de Desconto CFio HP": "Desc. Demanda",
        "Ajute de Desconto C.Fio HP": "Desc. Demanda",
        "Ajuste de Desc. C. Fio HP": "Desc. Demanda",
        "Ajuste Desc C.Fio HP": "Desc. Demanda",
        "Ajuste de Desconto C fio HP": "Desc. Demanda",
        "Ajuste Desconto C.Fio HP": "Desc. Demanda",
        "Ajuste de Desconto C Fio HP": "Desc. Demanda",
        "Ajuste de Desconto Cfio HP": "Desc. Demanda",
        "Ajuste de Desconto C.Fio HFP": "Desc. Demanda",
        "Ajuste de Desconto C. Fio HFP": "Desc. Demanda",
        "Ajuste de Desconto CFio HFP": "Desc. Demanda",
        "Ajute de Desconto C.Fio HFP": "Desc. Demanda",
        "Ajste de Desconto C.Fio HFP": "Desc. Demanda",
        "Ajuste de Desc. C. Fio HFP": "Desc. Demanda",
        "Ajuste Desc C.Fio HFP": "Desc. Demanda",
        "Ajuste de Desconto Cfio HFP": "Desc. Demanda",
        "Ajuste de Desconto C fio HFP": "Desc. Demanda",
        "DIF.FATUR.TUSD HOMOLOGACAO CCEE": "Desc. Demanda",
        "DIF.FATUR.TUSD-HOMOLOGACAO CCEE": "Desc. Demanda",
        "DIF.FATUR TUSD HOMOLOGACAO CCEE": "Desc. Demanda",
        "DIF. FATUR. TUSD HOMOLOG CCEE": "Desc. Demanda",
        "DIF FATUR TUSD HOMOLOGACAO CCEE": "Desc. Demanda",
        "DIF FATUR TUSD-HOMOLOGACAO CCEE": "Desc. Demanda",
        "DIF. FATURACAO TUSD CCEE": "Desc. Demanda",
        "DIFERENCA FATURAMENTO TUSD CCEE": "Desc. Demanda",
        "D1F.FATUR.TUSD-H0MOLOG.C0EE": "Desc. Demanda",
        "DIF.FATUR.T0SD HOMOLOG CCE": "Desc. Demanda",
        "DIF.FATUR.TUSD HOM0LOG CCE": "Desc. Demanda",
        "D1F FATURACAO TUSD CCEE": "Desc. Demanda",
        "AJUSTE DESC COMP FIO HP": "Desc. Demanda",
        "AJUSTE DESC. COMP. FIO HFP": "Desc. Demanda",
        "AJUSTE DESC. COMP. FIO HP": "Desc. Demanda",
        "AJUSTE DESCONTO COMP FIO HP": "Desc. Demanda",
        "DESC. COMPONENTE FIO HP": "Desc. Demanda",
        "DESC. COMPONENTE FIO HFP": "Desc. Demanda",
        "AJSTE DE DESCONTO C. FIO HP": "Desc. Demanda",
        "AJSTE DE DESCONTO C. FIO HFP": "Desc. Demanda",
        "DESC COMP. FIO HFP": "Desc. Demanda",
        "DESC COMP. FIO HP": "Desc. Demanda",
        "DESC COMP.FIO HP": "Desc. Demanda",
        "AJUSTEDESCONTOCFIOHFP": "Desc. Demanda",
        "DESC.COMPFIOHFP": "Desc. Demanda",
        "AJUSTEDESC.C.FIOHP": "Desc. Demanda",




#===================================================================Reativa===================================================================
        "UFER PONTA TE": "Reativa",
        "UFERFORA PONTA TE": "Reativa",
        "UFER PONTA  TE": "Reativa",
        "UFER PONTA TE ": "Reativa",
        "UFERF0RA PONTA TE": "Reativa",
        "UFER P0NTA TE": "Reativa",
        "UFER PONTA  TEE": "Reativa",
        "UFER PONTA-TE": "Reativa",
        "UFER PONTA.TE": "Reativa",
        "UFER F0RA PONTA TE": "Reativa",
        "UFFER PONTA TE": "Reativa",
        "Energia Reativa kWh HFP/unico": "Reativa",
        "energia reativa kwh hfp/unico": "Reativa",
        "Energia Reativa kWh HFP/Unico": "Reativa",
        "energia reativa kwh hfp/unico": "Reativa",
        "Energia Reativa kWh HP": "Reativa",
        "energia reativa kwh hp": "Reativa",
        "UFFER FORA PONTA TE": "Reativa",
        "UFER PONTA_Te": "Reativa",
        "UFER  PONTA TE": "Reativa",
        "UFER PONTA - TE": "Reativa",
        "UFER PONTA_Te": "Reativa",
        "UFER- PONTA TE": "Reativa",
        "UFERPONTA TE": "Reativa",
        "UFER P0NTA TE ": "Reativa",
        "UFER PONTA TE": "Reativa",
        "Energia Reativa HFP": "Reativa",
        "Energia Reativa HP": "Reativa",
        "UFERFORA PONTA TE": "Reativa",
        "UFER PONTA  TE": "Reativa",
        "UFER PONTA TE ": "Reativa",
        "UFERF0RA PONTA TE": "Reativa",
        "UFER P0NTA TE": "Reativa",
        "UFER PONTA  TEE": "Reativa",
        "UFER PONTA-TE": "Reativa",
        "UFER PONTA.TE": "Reativa",
        "UFER F0RA PONTA TE": "Reativa",
        "Consumo Reativo Excedente Fp": "Reativa",
        "Consumo Reativo Excedente Np": "Reativa",
        "UFER FP/unico (não Faturado)": "Reativa",
        "UFER FP/unico (nao Faturado)": "Reativa",
        "ufer fp/unico (nao faturado)": "Reativa",
        "ufer fp/unico (não faturado)": "Reativa",
        "ufer fp/unico (nao faturado)": "Reativa",
        "ufer fp unico (nao faturado)": "Reativa",
        "uferfp/unico (nao faturado)": "Reativa",
        "ufer fp/ünico (nao faturado)": "Reativa",
        "ufer fp/ünico (n40 faturado)": "Reativa",
        "ufer fp/unico (n4o faturado)": "Reativa",
        "ufer fp/unico nao faturado": "Reativa",
        "ufer fp unico nao faturado": "Reativa",
        "ufer fp/nico (nao faturado)": "Reativa",
        "ufer fp/unico (não fatvrado)": "Reativa",
        "ufer fp/un1co (n4o faturado)": "Reativa",
        "ufer fp/un1co (nao faturado)": "Reativa",
        "ufer fp/unico (n faturado)": "Reativa",
        "ufer fp/unico (nao faturad0)": "Reativa",
        "ufer fp/ünico nao faturado": "Reativa",
        "uferfp unico nao faturado": "Reativa",
        "ufer fp/unico (nao fat)": "Reativa",
        "ufer fp unico (não fat)": "Reativa",
        "UFER FP/ünico (näo Faturado)": "Reativa",
        "UFER FP/un1co (n40 Faturado)": "Reativa",
        "UFER FP/unico (n Faturado)": "Reativa",
        "UFER FP unico nao Faturado": "Reativa",
        "UFER FP/unico nao Faturado": "Reativa",
        "UFER FP/nico (nao Faturado)": "Reativa",
        "Energia Reativa kWh HFP/Unico": "Reativa",
        "Energia Reativa kWh HFP/unico": "Reativa",
        "Energia Reativa kWh HFP/unico": "Reativa",
        "Energia Reativa kWh HFP/unico": "Reativa",
        "Energia Reativa kWh HFP": "Reativa",
        "Energia Reativa kWhHFP/Unico": "Reativa",
        "Energia Reativa kWh HFP Unico": "Reativa",
        "Energia Reativa kwh hfp unico": "Reativa",
        "Energia Reativa kwh hfp/nico": "Reativa",
        "Energia Reativa kwh HFP/ni co": "Reativa",
        "Energia Reativa kwhHFP/nico": "Reativa",
        "Energia Reativa kwhhfp": "Reativa",
        "Energia Reativa kWh HFPUnico": "Reativa",
        "Energia Reatíva kWh HFP/Unico": "Reativa",
        "Energia Reativa kWh HFP/U1nico": "Reativa",
        "Energia Reativa kWh HFP/Un1co": "Reativa",
        "Energia Reativa kWh HFP Ünico": "Reativa",
        "Energia Reativa kWh HFP ùnico": "Reativa",
        "Energia Reativa kWh HP": "Reativa",
        "Energia Reativa kwh HP": "Reativa",
        "Energia Reativa kWhHP": "Reativa",
        "Energia Reativa kwH HP": "Reativa",
        "Energia Reatíva kWh HP": "Reativa",
        "Energia Reativa kWh H P": "Reativa",
        "Energia Reativa KWh HP": "Reativa",
        "Energia Reativa kwh h p": "Reativa",
        "EnergiaReativa kWh HP": "Reativa",
        "Energia Reativa kwHP": "Reativa",
        "Energia Reativa kwh": "Reativa",
        "UFER FP UNICO": "Reativa",
        "UFER FP uNICO": "Reativa",
        "UFER FP/UNICO": "Reativa",
        "UFER FP/UN1CO": "Reativa",
        "UFERFP/UNICO": "Reativa",
        "UFER FP UNICO NAO FATURADO": "Reativa",
        "UFER FP UNICO (NAO FATURADO)": "Reativa",
        "UFER FP uNICO (NÃO FATURADO)": "Reativa",
        "UFERFP UNICO (NAO FATURADO)": "Reativa",
        "UFER FPUNICO (NAO FATURADO)": "Reativa",
        "UFER FP/NICO (NAO FATURADO)": "Reativa",
        "UFER FP/NlCO (NAO FATURADO)": "Reativa",
        "UFER FP/ÜNICO (NAO FATURADO)": "Reativa",
        "UFER FP/UNIC0 (NAO FATURADO)": "Reativa",
        "UFER FP/UN1CO NAO FATURADO": "Reativa",
        "ENERGIA REATIVA HFP": "Reativa",
        "ENERGIA REATIVA HP": "Reativa",
        "ENERGIA REATIVA KWH HFP": "Reativa",
        "ENERGIA REATIVA KWH HP": "Reativa",
        "ENERGIA REATIVA KWHHFP": "Reativa",
        "ENERGIA REATIVA KWHHP": "Reativa",
        "ENERGIAREATIVA KWH HP": "Reativa",
        "ENERGIA REATIVA KWHP": "Reativa",
        "ENERGIA REATIVA KWH HFP UNICO": "Reativa",
        "ENERGIA REATIVA KWH HFP/UNICO": "Reativa",
        "ENERGIA REATIVA KWH HFP/uNICO": "Reativa",
        "ENERGIA REATIVA KWH HFP/NICO": "Reativa",
        "ENERGIA REATIVA KWH HFP/ÜNICO": "Reativa",
        "ENERGIA REATIVA KWH HFP/UN1CO": "Reativa",
        "ENERGIA REATIVA KWH HFP/U1NICO": "Reativa",
        "ENERGIA REATIVA KWH HFP/UNCO": "Reativa",
        "ENERGIA REATIVA KWH H P": "Reativa",
        "ENERGIA REATIVA KWH HFPUNICO": "Reativa",
        "ENERGIA REATIVA KW HFP": "Reativa",
        "ENERGIA REATIVA KWH": "Reativa",
        "ENERGIAREATIVA KWH": "Reativa",
        "ENERGIA REATIVA HP ": "Reativa",
        "UFER PONTA TEE": "Reativa",
        "UFER FORA PONTA TE": "Reativa",
        "UFER FORA P0NTA TE": "Reativa",
        "UFER FORA PONTA-TE": "Reativa",
        "UFER FORA PONTA.TE": "Reativa",
        "UFER-FORA PONTA TE": "Reativa",
        "UFER PONTA TE (NÃO FATURADO)": "Reativa",



#===================================================================Cont Publica===================================================================
        "CIP-ILUM PUB PREF MUNICIPAL": "Contb. Publica",
        "CIP - ILUM PUB PREF MUNICIPAL": "Contb. Publica",
        "COSIP.SAOPAULO-MUNICIPAL": "Contb. Publica",
        "CIP-BARUERI-MUNICIPAL": "Contb. Publica",
        "CIP-CAJAMAR-MUNICIPAL": "Contb. Publica",
        "CIP-ILUM PUB PREF MUNICIPALCONTRIB PUB": "Contb. Publica",
        "CIP-SAOPAULO-MUNICIPAL": "Contb. Publica",
        "CIP-ILUM PUB PREF MUNICIPAL-SAO PAULO": "Contb. Publica",
        "CIP_ILUM_PUB_MUNICIPAL": "Contb. Publica",
        "Contrib llum Publica Municipal": "Contb. Publica",
        "Contrib Ilum Publica Municipal": "Contb. Publica",
        "Contrib Ilum Publica Municipal": "Contb. Publica",
        "Contrib IlumPublica Municipal": "Contb. Publica",
        "Contrib IlumPublicaMunicipal": "Contb. Publica",
        "Contrib llumPublica Municipal": "Contb. Publica",
        "Contrib llumPublicaMunicipal": "Contb. Publica",
                                 "Contrib Ilum Publica Municipal": "Contb. Publica",
    "contrib ilum publica municipal": "Contb. Publica",
    "Contrib Ilum Publica Municipal": "Contb. Publica",
    "contrib ilum publica municipal": "Contb. Publica",
        "CIPILUM PUB PREF MUNICIPAL": "Contb. Publica",
        "CIP ILUM PUB PREF MUNICIPAL": "Contb. Publica",
        "C1P - ILUM PUB PREF MUNICIPAL": "Contb. Publica",
        "C1P-ILUM PUB PREF MUNICIPAL": "Contb. Publica",
        "CIP ILUM PUBL PREF MUNICIPAL": "Contb. Publica",
        "COSIP SAOPAULO MUNICIPAL": "Contb. Publica",
        "COSIP SAO PAULO MUNICIPAL": "Contb. Publica",
        "C0SIP.SAOPAULO-MUNICIPAL": "Contb. Publica",
        "CIP-BARUERI MUNICIPAL": "Contb. Publica",
        "CIP CAJAMAR MUNICIPAL": "Contb. Publica",
        "CIP ILUM PUB MUNICIPAL": "Contb. Pblica",
        "CIP-ILUM PUB PREF MUNICIPAL CONTRIB PUBLICA": "Contb. Publica",
        "CIP ILUM PUBLICA MUNICIPAL": "Contb. Publica",
        "CIP-ILUM PUBLICA MUNICIPAL": "Contb. Publica",
        "Contrib llum Publica Municpal": "Contb. Publica",
        "Contrib llum Pubica Municipal": "Contb. Publica",
        "Contrib Ilum Publica Municpal": "Contb. Publica",
        "Contribuçao Ilum Publica": "Contb. Publica",
        "Contribuiçao Ilum Publica": "Contb. Publica",
        "Contribuiçâo llum Publica Municipal": "Contb. Publica",
        "Cont llum Publica": "Contb. Publica",
        "Contrib IlumPublicaMunicipal": "Contb. Publica",
        "Contr llumPublicaMunicipal": "Contb. Publica",
        "Contribuicao Ilum Publica": "Contb. Publica",
        "Contribuicao IlumPublica": "Contb. Publica",
        "Contribuição Ilum Publica Municipal": "Contb. Publica",
        "Cont Ilum Publica": "Contb. Publica",
        "C0ntrib Ilum Publica": "Contb. Publica",
        "Contrib llum Püblica Municipal": "Contb. Publica",
        "Contribuição Iluminação Publica": "Contb. Publica",
        "Cont llum Publica": "Contb. Publica",
        "Contribuição Publica": "Contb. Publica",
        "Contrib llum Publica": "Contb. Publica",
        "Contrib IlumMunicipal": "Contb. Publica",
        "Contrib llum Püblica Municipal": "Contb. Publica",
        "Contrib Ilum Publica Municipal": "Contb. Publica",
        "Contrib Ilum Publica Municipal": "Contb. Publica",
        "Contrib IlumPublica Municipal": "Contb. Publica",
        "Contrib llumPublica Municipal": "Contb. Publica",
        "Contrib IlumPüblica Municipal": "Contb. Publica",
        "Contrib llum Publica Municipal": "Contb. Publica",
        "Contrib llum Püblca Municipal": "Contb. Publica",
        "Contrib llum Publica Municpal": "Contb. Publica",
        "Contrib llum Püblca Municipal": "Contb. Publica",
        "Contrib llum Publica Municpal": "Contb. Publica",
        "Contrib llum Püblica Municipal": "Contb. Publica",
        "Contrib llum Publica Municpal": "Contb. Publica",
        "Contrib Ilum Publica Municipal": "Contb. Publica",
        "Contrib Ilum Publica Municipal": "Contb. Publica",
        "Contrib IlumPublica Municipal": "Contb. Publica",
        "Contrib llumPublica Municipal": "Contb. Publica",
        "Contrib IlumPüblica Municipal": "Contb. Publica",
        "Contrib llum Publica Municipal": "Contb. Publica",
        "Contrib llum Püblca Municipal": "Contb. Publica",
}


# Funções auxiliares
def normalizar_texto(texto):
    if not isinstance(texto, str):
        return ""
    texto = texto.lower().strip()
    texto = unicodedata.normalize('NFKD', texto).encode('ASCII', 'ignore').decode('ASCII')
    texto = re.sub(r'[^a-zA-Z0-9]', '', texto)
    return texto

def calcular_custo_api(response):
    try:
        usage = response.usage
        input_tokens = usage.prompt_tokens
        output_tokens = usage.completion_tokens
        
        # Preços para GPT-3.5-turbo (em USD por 1K tokens)
        preco_input = 0.0005
        preco_output = 0.0015
        
        custo = (input_tokens * preco_input + output_tokens * preco_output) / 1000
        return custo
    except:
        return 0

def extrair_texto(pdf_path_ou_bytes):
    try:
        if isinstance(pdf_path_ou_bytes, bytes):
            # Caso seja um objeto de bytes (upload via Streamlit)
            with fitz.open(stream=pdf_path_ou_bytes, filetype="pdf") as doc:
                texto = ""
                for pagina in doc:
                    texto += pagina.get_text()
                return texto
        else:
            # Caso seja um caminho de arquivo
            with fitz.open(pdf_path_ou_bytes) as doc:
                texto = ""
                for pagina in doc:
                    texto += pagina.get_text()
                return texto
    except Exception as e:
        st.error(f"Erro ao extrair texto do PDF: {e}")
        # Tentar com pikepdf para PDFs protegidos
        try:
            with pikepdf.open(pdf_path_ou_bytes, password="") as pdf:
                with fitz.open(stream=pdf.save(), filetype="pdf") as doc:
                    texto = ""
                    for pagina in doc:
                        texto += pagina.get_text()
                    return texto
        except Exception as e2:
            # Tentar com senhas comuns
            senhas_tentativas = ["33042", "56993", "60869", "08902", "3304", "5699", "6086", "0890"]
            for senha in senhas_tentativas:
                try:
                    with pikepdf.open(pdf_path_ou_bytes, password=senha) as pdf:
                        with fitz.open(stream=pdf.save(), filetype="pdf") as doc:
                            texto = ""
                            for pagina in doc:
                                texto += pagina.get_text()
                            return texto
                except:
                    continue
            
            st.error(f"Não foi possível extrair o texto do PDF. O arquivo pode estar corrompido ou protegido.")
            return ""

def extrair_trecho_itens_fatura(texto_pdf):
    padrao_inicio = r"(?i)itens\s+da\s+fatura"
    padrao_fim = r"(?i)(total\s+da\s+fatura|total\s+a\s+pagar|total\s+geral|valor\s+total|total\s+da\s+nota)"
    
    match_inicio = re.search(padrao_inicio, texto_pdf)
    if not match_inicio:
        return ""
    
    inicio = match_inicio.start()
    trecho = texto_pdf[inicio:]
    
    match_fim = re.search(padrao_fim, trecho)
    fim = len(trecho)
    if match_fim:
        idx = match_fim.end()
        # Pegar mais algumas linhas após o "Total"
        linhas_apos_total = trecho[idx:].split('\n')[:5]
        idx += sum(len(linha) + 1 for linha in linhas_apos_total)
        if 0 < idx < fim:
            fim = idx
    return trecho[:fim]

def corrigir_negativo_final(valor):
    if isinstance(valor, str) and valor.strip().endswith('-'):
        valor_corrigido = '-' + valor.strip()[:-1]
        return valor_corrigido
    return valor

def formatar_numero_brasileiro(x):
    try:
        num = float(str(x).replace(",", "."))
        return f"{num:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except:
        return x

# NOVA FUNÇÃO: Extrair Número de Instalação e Referência do PDF
def extrair_info_cabecalho(texto_pdf, client):
    # Limitar a 500 palavras para melhorar a performance
    texto_limitado = " ".join(texto_pdf.split()[:150])

    prompt_extracao_cabecalho = """
    Você é um assistente especializado em extrair informações de documentos.

    Extraia APENAS as seguintes informações da conta de energia elétrica:
    1. Número de Instalação (pode aparecer como "Nº da Instalação", "Instalação", "Nº Instalação", "Unid. Consumidora", "UC", "Código do Cliente", etc.)
    2. Referência (geralmente um mês/ano, como "01/2025", "JAN/2025", "JANEIRO/2025", etc.)
    3.A REFERECIA SEMPRE E UM VALOR NUMERICO!!!

    Retorne apenas um objeto JSON com os seguintes campos:
    {
        "N_Instalacao": "valor extraído",
        "Referencia": "valor extraído"
    }

    Se alguma informação não for encontrada, use string vazia.
    Não inclua explicações, apenas o JSON.
    """

    try:
        response = client.chat.completions.create(
            model="gpt-3.5-turbo-0125",
            messages=[
                {"role": "system", "content": prompt_extracao_cabecalho},
                {"role": "user", "content": texto_limitado}
            ],
            temperature=0
        )

        custo = calcular_custo_api(response)
        resposta = response.choices[0].message.content.strip()

        # Limpar qualquer formatação extra
        if resposta.startswith("```json"):
            resposta = resposta.replace("```json", "").replace("```", "").strip()

        try:
            dados = json.loads(resposta)
            return dados, custo
        except:
            # Se falhar ao fazer parse do JSON, retorna valores vazios
            return {"N_Instalacao": "", "Referencia": ""}, custo

    except Exception as e:
        st.error(f"Erro ao extrair informações de cabeçalho: {e}")
        return {"N_Instalacao": "", "Referencia": ""}, 0

# Interface Streamlit
st.markdown('<h1 class="main-header">⚡ PDFator - Extrator de Faturas de Energia</h1>', unsafe_allow_html=True)
st.markdown('<p class="sub-header">Extraia e categorize informações de suas faturas de energia em poucos cliques.</p>', unsafe_allow_html=True)

# Seção para API Key
with st.expander("🔑 Insira sua chave da API OpenAI:", expanded=True):
    api_key = st.text_input("Chave da API OpenAI", type="password", help="Sua chave não será salva no código")

# Inicializar variáveis de estado da sessão
if 'custo_total_usd' not in st.session_state:
    st.session_state.custo_total_usd = 0
    
if 'pdfs_processados' not in st.session_state:
    st.session_state.pdfs_processados = 0

# Estatísticas
col1, col2 = st.columns(2)
with col1:
    st.metric("📄 PDFs processados", st.session_state.pdfs_processados)
with col2:
    custo_brl = st.session_state.custo_total_usd * 5.67  # Taxa de conversão aproximada
    st.metric("💰 Custo estimado", f"${st.session_state.custo_total_usd:.6f} USD ≈ R${custo_brl:.2f}")

# Upload de arquivos
with st.expander("📂 Selecione seus arquivos PDF (máximo 80 arquivos):", expanded=True):
    uploaded_files = st.file_uploader("Arraste e solte arquivos aqui", type="pdf", accept_multiple_files=True, help="Limite 200MB por arquivo • PDF")

# Botão de processamento
if st.button("Processar PDFs", type="primary", disabled=not api_key or not uploaded_files):
    if not api_key:
        st.error("Por favor, insira sua chave da API OpenAI.")
    elif not uploaded_files:
        st.error("Por favor, selecione pelo menos um arquivo PDF.")
    else:
        # Inicializar cliente OpenAI
        client = OpenAI(api_key=api_key)
        
        # Inicializar novo Excel e variável de custo
        wb = openpyxl.Workbook()
        custo_total_usd = 0
        
        # Criar aba de resumo
        if 'Sheet' in wb.sheetnames:
            ws_resumo = wb['Sheet']
            ws_resumo.title = "Resumo"
        else:
            ws_resumo = wb.create_sheet(title="Resumo")
        
        ws_resumo.sheet_view.showGridLines = False
        
        # Adicionar cabeçalho na aba de resumo
        cabecalhos_resumo = ["Nome Do PDF", "N Instalacao", "Referencia", "Encargo", "Desc. Encargo", "Reativa", "Demanda", "Desc. Demanda", "Contrib. Publica"]
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        
        for col, header in enumerate(cabecalhos_resumo, start=1):
            cell = ws_resumo.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # Lista para armazenar dados do resumo
        dados_resumo = []
        
        # Barra de progresso
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        # === LAÇO PRINCIPAL: Processar cada PDF ===
        for idx_pdf, uploaded_file in enumerate(uploaded_files[:80]):  # Limitar a 80 arquivos
            try:
                # Atualizar progresso
                progress = (idx_pdf + 1) / min(len(uploaded_files), 80)
                progress_bar.progress(progress)
                status_text.text(f"🔎 Processando arquivo {idx_pdf + 1}/{min(len(uploaded_files), 80)}: {uploaded_file.name}...")
                
                nome_arquivo = uploaded_file.name.rsplit(".", 1)[0][:31]
                
                # Ler o conteúdo do arquivo
                pdf_bytes = uploaded_file.read()
                texto_completo = extrair_texto(pdf_bytes)
                
                # Extrair informações de cabeçalho (N. Instalação e Referência)
                info_cabecalho, custo_cabecalho = extrair_info_cabecalho(texto_completo, client)
                custo_total_usd += custo_cabecalho
                
                texto_tabela = extrair_trecho_itens_fatura(texto_completo)
                
                if not texto_tabela.strip():
                    st.warning(f"🚨 Seção Itens da Fatura não encontrada no arquivo {uploaded_file.name}!")
                    continue
                
                prompt_extracao = """
                Você é um extrator especializado em identificar e estruturar tabelas extraídas de documentos PDF.

                Instruções:
                - Você receberá o conteúdo de um trecho de um PDF.
                - Localize a tabela que se inicia com o título "Itens da Fatura" e processe até o último valor numérico listado.
                - Cada linha da tabela deve ser transformada em um objeto JSON com os seguintes campos obrigatórios:
                  ["Item", "Unidade", "Quantidade", "Preço Unitário (R$)", "Valor (R$)", "PIS/COFINS", "Base de Cálculo ICMS", "Alíquota ICMS", "ICMS (R$)", "Tarifa Unitária"].

                Regras de extração:
                - Ignore qualquer linha de uma estrutracao de tabela que comece com "Medidor" e nao deixe isso afetar a extracao para outras linhas.
                - A linha que contém o texto "TOTAL" deve ser separada como um objeto independente.
                - Não mescle o conteúdo da linha "TOTAL" com nenhuma outra linha da tabela.
                - Se uma linha contiver apenas um nome e números, associe o maior número ao campo "Valor (R$)".
                - Não infira (não invente) valores ou informações ausentes.
                - Campos vazios devem ser preenchidos com a string vazia: `""`.
                - Extraia corretamente números negativos, mesmo quando o sinal de negativo estiver **posicionado após o número** (exemplo: 32123,23- deve ser interpretado como -32123,23).
                - É essencial que você extraia TODAS as linhas da tabela.
                Regra especial:
                - Se encontrar a linha "Contrib Ilum Publica Municipal":
                    - Preencher apenas o campo "Valor (R$)" com o valor encontrado.
                    - Deixar os campos "Unidade", "Quantidade", "Preço Unitário (R$)", "PIS/COFINS", "Base de Cálculo ICMS", "Alíquota ICMS", "ICMS (R$)", "Tarifa Unitária" como string vazia `""`.

                Caso a tabela não seja encontrada:
                - Retorne exatamente: ERRO: tabela nao encontrada.

                Importante:
                - Seja extremamente rigoroso na separação de colunas e na associação dos números aos campos corretos.
                """
                
                response = client.chat.completions.create(
                    model="gpt-3.5-turbo-0125",
                    messages=[
                        {"role": "system", "content": prompt_extracao},
                        {"role": "user", "content": texto_tabela}
                    ],
                    temperature=0,
                    max_tokens=4096
                )
                
                custo_total_usd += calcular_custo_api(response)
                
                resposta = response.choices[0].message.content.strip()
                if resposta.startswith("```json"):
                    resposta = resposta.replace("```json", "").replace("```", "").strip()
                if resposta.lower().startswith("erro"):
                    st.warning(f"🚨 Tabela não encontrada no arquivo {uploaded_file.name}!")
                    continue
                
                tabela = pd.DataFrame(json.loads(resposta))
                resumo = {"Encargo": 0, "Desc. Encargo": 0, "Demanda": 0, "Desc. Demanda": 0, "Reativa": 0, "Contb. Publica": 0, "Outros": 0}
                
                for _, row in tabela.iterrows():
                    item = str(row["Item"]).strip().lower()
                    valor_corrigido = corrigir_negativo_final(str(row["Valor (R$)"]))
                    valor = pd.to_numeric(valor_corrigido.replace(".", "").replace(",", "."), errors='coerce')
                    if pd.isna(valor) or item == "":
                        continue
                    if any(termo in item for termo in ["total", "subtotal faturamento", "subtotal outros"]):
                        continue
                    
                    tipo = "Outros"
                    for chave in mapeamento_tipo_cobranca:
                        if normalizar_texto(chave) in normalizar_texto(item) or normalizar_texto(item) in normalizar_texto(chave):
                            tipo = mapeamento_tipo_cobranca[chave]
                            break
                    resumo[tipo] += valor
                
                total_resumo = sum(resumo.values())
                
                # Adicionar à lista de dados para a aba resumo
                dados_resumo.append({
                    "Nome Do PDF": nome_arquivo,
                    "N Instalacao": info_cabecalho.get("N_Instalacao", ""),
                    "Referencia": info_cabecalho.get("Referencia", ""),
                    "Encargo": resumo.get("Encargo", 0),
                    "Desc. Encargo": resumo.get("Desc. Encargo", 0),
                    "Reativa": resumo.get("Reativa", 0),
                    "Demanda": resumo.get("Demanda", 0),
                    "Desc. Demanda": resumo.get("Desc. Demanda", 0),
                    "Contrib. Publica": resumo.get("Contb. Publica", 0)
                })
                
                colunas_numericas = ["Quantidade", "Preço Unitário (R$)", "Valor (R$)", "PIS/COFINS", "Base de Cálculo ICMS", "Alíquota ICMS", "ICMS (R$)", "Tarifa Unitária"]
                for coluna in colunas_numericas:
                    if coluna in tabela.columns:
                        tabela[coluna] = tabela[coluna].apply(corrigir_negativo_final).apply(formatar_numero_brasileiro)
                
                # Criar nova aba para este PDF
                ws = wb.create_sheet(title=nome_arquivo)
                ws.sheet_view.showGridLines = False
                
                start_row = 3
                start_col = 2
                
                # Adicionar cabeçalho com estilo
                for j, col_name in enumerate(tabela.columns, start=start_col):
                    cell = ws.cell(row=start_row, column=j, value=col_name)
                    cell.font = header_font
                    cell.fill = header_fill
                
                # Adicionar dados da tabela
                for i, row in enumerate(tabela.values, start=start_row + 1):
                    for j, value in enumerate(row, start=start_col):
                        ws.cell(row=i, column=j, value=value)
                
                # Ajustar largura das colunas
                for col in ws.columns:
                    max_length = 0
                    col_letter = get_column_letter(col[0].column)
                    for cell in col:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    ws.column_dimensions[col_letter].width = max(max_length + 2, 10)
                
                # Adicionar resumo na aba do PDF
                resumo_row_start = start_row + len(tabela) + 5
                
                ws.cell(row=resumo_row_start, column=start_col, value="Tipo de Cobrança").font = header_font
                ws.cell(row=resumo_row_start, column=start_col, value="Tipo de Cobrança").fill = header_fill
                ws.cell(row=resumo_row_start, column=start_col + 1, value="Valor Total (R$)").font = header_font
                ws.cell(row=resumo_row_start, column=start_col + 1, value="Valor Total (R$)").fill = header_fill
                
                tipos = ["Encargo", "Desc. Encargo", "Demanda", "Desc. Demanda", "Reativa", "Contb. Publica", "Outros", "Total"]
                for idx_tipo, tipo in enumerate(tipos):
                    valor_tipo = resumo.get(tipo, 0) if tipo != "Total" else total_resumo
                    ws.cell(row=resumo_row_start + idx_tipo + 1, column=start_col, value=tipo)
                    ws.cell(row=resumo_row_start + idx_tipo + 1, column=start_col + 1, value=formatar_numero_brasileiro(round(valor_tipo, 2)))
                
                # Atualizar contador de PDFs processados
                st.session_state.pdfs_processados += 1
                
            except Exception as e:
                st.error(f"Erro ao processar {uploaded_file.name}: {e}")
        
        # Preencher a aba de resumo com os dados coletados
        for i, dados in enumerate(dados_resumo, start=2):
            for j, col in enumerate(cabecalhos_resumo, start=1):
                valor = dados.get(col, "")
                # Formatar valores numéricos
                if isinstance(valor, (int, float)) and col not in ["Nome Do PDF", "N Instalacao", "Referencia"]:
                    ws_resumo.cell(row=i, column=j, value=formatar_numero_brasileiro(round(valor, 2)))
                else:
                    ws_resumo.cell(row=i, column=j, value=valor)
        
        # Ajustar largura das colunas na aba de resumo
        for col in ws_resumo.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws_resumo.column_dimensions[col_letter].width = max(max_length + 2, 12)
        
        # Remover a aba Sheet padrão se existir e não tiver sido usada
        if 'Sheet' in wb.sheetnames and 'Sheet' != ws_resumo.title:
            del wb['Sheet']
        
        # Salvar o arquivo Excel em memória
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # Atualizar custo total
        st.session_state.custo_total_usd += custo_total_usd
        
        # Oferecer download do arquivo Excel
        st.success("✅ Todos os arquivos foram processados!")
        st.download_button(
            label="📥 Baixar Relatório Excel",
            data=excel_buffer,
            file_name="extracao_tabela_final.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
        # Atualizar estatísticas
        custo_brl = st.session_state.custo_total_usd * 5.67  # Taxa de conversão aproximada
        st.metric("💰 Custo total estimado", f"${st.session_state.custo_total_usd:.6f} USD ≈ R${custo_brl:.2f}")

# Rodapé
st.markdown('<div class="footer">📝 PDFator - Extrator de Faturas de Energia | 2025</div>', unsafe_allow_html=True)
